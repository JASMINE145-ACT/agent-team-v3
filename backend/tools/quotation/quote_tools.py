# 报价 Agent 工具：从原始报价单 Excel 提取「第2行到 Total Excluding PPN不含税总价 所在行之前」的报价数据，供 LLM 使用

from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any, List

# 边界行标识（用户指定：报价数据列到该行为止）
TOTAL_ROW_MARKER = "Total Excluding PPN不含税总价"

# 询价列表头关键词（用于自动识别列）
NAME_COL_KEYWORDS = ["询价货物名称", "Nama Permintaan Barang", "nama permintaan"]
SPEC_COL_KEYWORDS = ["询价规格型号", "Spesifikasi dan Model Permintaan Barang", "Spesifikasi"]
QTY_COL_KEYWORDS = ["Jumlah", "数量", "jumlah", "Quantity"]


def _cell_value(cell) -> str:
    v = getattr(cell, "value", None)
    if v is None:
        return ""
    return str(v).strip()


def extract_quotation_data(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    从原始报价单 Excel 提取报价数据：从第 2 行起到「Total Excluding PPN不含税总价」所在行的上一行止。

    - 第 1 行视为表头。
    - 数据区：第 2 行 ～ 第一个包含 TOTAL_ROW_MARKER 的行的上一行。
    - 支持 .xlsx / .xlsm；.xls 需调用方先转为 xlsx 或在此用 pandas 扩展。

    Returns:
        {"success": bool, "result": str, "error": str | None, "rows_count": int}
        result 为 Markdown 表格或 JSON 文本，便于 LLM 阅读。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm；.xls 请先另存为 xlsx", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        # 先收集所有行（read_only 下 iter_rows 只能遍历一次，先找边界行）
        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)

            # 任意单元格包含边界标识即视为「总价行」，数据区在其上一行结束
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "result": "", "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "result": "表为空，无报价数据。", "error": None, "rows_count": 0}

    # 表头：第 1 行（索引 0）
    header = all_rows[0]
    # 数据行：第 2 行到「Total Excluding PPN」的上一行（不包含总价行）
    if total_row_1based is not None and total_row_1based >= 2:
        data_rows = all_rows[1 : total_row_1based - 1]
    else:
        data_rows = all_rows[1:]

    if not data_rows:
        return {"success": True, "result": "未找到数据行（或仅含总价行）。", "error": None, "rows_count": 0}

    # 输出为 Markdown 表格，便于 LLM 理解
    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(header), max(len(r) for r in data_rows), 1)
    header_padded = header + [""] * (col_count - len(header))
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    lines = ["| " + " | ".join(escape_md(h) for h in header_padded) + " |", sep]
    for r in data_rows:
        r_padded = r + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(c) for c in r_padded) + " |")

    result_text = "报价数据（第2行至「Total Excluding PPN不含税总价」上一行）：\n\n" + "\n".join(lines)
    return {
        "success": True,
        "result": result_text,
        "error": None,
        "rows_count": len(data_rows),
    }


def _find_col_by_header(header: List[str], keywords_list: List[str]) -> int:
    """按表头关键词查找列索引，返回 -1 表示未找到。"""
    for i, cell in enumerate(header):
        val = (str(cell or "").strip()).lower()
        for kw in keywords_list:
            if kw.lower() in val:
                return i
    return -1


def _extract_inquiry_items_smart_fallback(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 200,
) -> dict[str, Any]:
    """
    普适解析 fallback：不依赖 TOTAL_ROW_MARKER，读取工作表前 max_rows 行，
    在前 3 行中按关键词识别名称/规格/数量列，构建 items。供 extract_inquiry_items 在主逻辑失败或无数据时调用。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not rows or len(rows) < 2:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    name_col = spec_col = qty_col = -1
    header_row_idx = 0
    for idx, header_row in enumerate(rows[:3]):
        nc = _find_col_by_header(header_row, NAME_COL_KEYWORDS)
        sc = _find_col_by_header(header_row, SPEC_COL_KEYWORDS)
        qc = _find_col_by_header(header_row, QTY_COL_KEYWORDS)
        if nc >= 0:
            name_col, spec_col = nc, sc
            if qc >= 0:
                qty_col = qc
            header_row_idx = idx
            break

    if name_col < 0:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    data_rows = rows[header_row_idx + 1:]
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = header_row_idx + 2 + i
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
        })

    return {"success": True, "items": items, "error": None, "rows_count": len(items)}


def extract_inquiry_items(
    file_path: str,
    sheet_name: str | None = None,
    col_mapping: dict | None = None,
) -> dict[str, Any]:
    """
    提取「询价货物名称」「询价规格型号」两列，输出供库存查询 Agent 用的列表。

    - 复用 data 区域识别逻辑（第2行～Total Excluding PPN 上一行）
    - 列识别：按表头匹配，或通过 col_mapping 指定 {name_col: int, spec_col: int}
    - keywords = product_name + " " + specification（空规格则仅名称）

    Returns:
        {"success": bool, "items": [...], "error": str | None, "rows_count": int}
        items: [{"row": 1, "product_name": "", "specification": "", "keywords": ""}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "items": [], "error": "请安装 openpyxl: pip install openpyxl", "rows_count": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "items": [], "error": f"文件不存在: {path}", "rows_count": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "items": [], "error": "仅支持 .xlsx / .xlsm", "rows_count": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "items": [], "error": f"打开 Excel 失败: {e}", "rows_count": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        all_rows: List[List[str]] = []
        total_row_1based: int | None = None

        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            cells = [_cell_value(c) for c in row]
            all_rows.append(cells)
            for c in cells:
                if TOTAL_ROW_MARKER in (c or ""):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break

        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "items": [], "error": str(e), "rows_count": 0}

    if not all_rows:
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # 确定表头行与列索引（部分模板首行为标题，次行为表头）
    if col_mapping:
        name_col = col_mapping.get("name_col", col_mapping.get("product_name_col", -1))
        spec_col = col_mapping.get("spec_col", col_mapping.get("specification_col", -1))
        qty_col = col_mapping.get("qty_col", col_mapping.get("quantity_col", -1))
        header_row_idx = 0
    else:
        name_col = spec_col = qty_col = -1
        header_row_idx = 0
        for idx, header_row in enumerate(all_rows[:3]):
            nc = _find_col_by_header(header_row, NAME_COL_KEYWORDS)
            sc = _find_col_by_header(header_row, SPEC_COL_KEYWORDS)
            qc = _find_col_by_header(header_row, QTY_COL_KEYWORDS)
            if nc >= 0:
                name_col, spec_col = nc, sc
                if qc >= 0:
                    qty_col = qc
                header_row_idx = idx
                break

    if name_col < 0:
        # Fallback：用普适解析（不依赖 Total Excluding PPN 与固定表头）再尝试识别列
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            fallback["error"] = None
            return fallback
        return {"success": False, "items": [], "error": "未找到询价货物名称列，请检查表头或提供 col_mapping", "rows_count": 0}

    # 数据行从表头下一行起，到 Total Excluding PPN 上一行
    data_start = header_row_idx + 1
    if total_row_1based is not None and total_row_1based >= 2:
        data_end = total_row_1based - 1
    else:
        data_end = len(all_rows)
    data_rows = all_rows[data_start:data_end]

    if not data_rows:
        fallback = _extract_inquiry_items_smart_fallback(file_path, sheet_name)
        if fallback.get("items"):
            fallback["_fallback_used"] = True
            return fallback
        return {"success": True, "items": [], "error": None, "rows_count": 0}

    # spec_col 可为 -1，表示无规格列；qty_col 可为 -1，表示无数量列
    items: List[dict] = []
    for i, row_cells in enumerate(data_rows):
        row_num = data_start + 1 + i  # Excel 行号 1-based
        product_name = (row_cells[name_col] if name_col < len(row_cells) else "").strip()
        specification = (row_cells[spec_col] if spec_col >= 0 and spec_col < len(row_cells) else "").strip()
        keywords = f"{product_name} {specification}".strip() if specification else product_name
        if not keywords:
            continue
        # 需求数量 qty：用于库存比对
        qty_val = 0
        if qty_col >= 0 and qty_col < len(row_cells):
            try:
                v = row_cells[qty_col]
                if v is not None and str(v).strip():
                    qty_val = int(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass
        items.append({
            "row": row_num,
            "product_name": product_name,
            "specification": specification,
            "keywords": keywords,
            "qty": qty_val,
        })

    return {
        "success": True,
        "items": items,
        "error": None,
        "rows_count": len(items),
    }


# ---------------------------------------------------------------------------
# 普适性 Excel 工具（不依赖报价单结构，任意 Excel 可用）
# ---------------------------------------------------------------------------

def parse_excel_smart(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 500,
) -> dict[str, Any]:
    """
    【普适性】智能解析任意 Excel：自动读取指定工作表的所有单元格（或前 max_rows 行），
    零硬编码列/行，适合多表头、合并单元格、不规则布局。返回 Markdown 表格便于 LLM 理解。

    Returns:
        {"success": bool, "result": str, "error": str | None, "sheet_name": str, "rows_read": int}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "sheet_name": "", "rows_read": 0}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "sheet_name": "", "rows_read": 0}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "result": "", "error": "仅支持 .xlsx / .xlsm", "sheet_name": "", "rows_read": 0}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "result": "", "error": f"打开 Excel 失败: {e}", "sheet_name": "", "rows_read": 0}

    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            used_sheet = sheet_name
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
            used_sheet = ws.title

        rows: List[List[str]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([str(c).strip() if c is not None else "" for c in (row or [])])
        wb.close()
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"success": False, "result": "", "error": str(e), "sheet_name": used_sheet if "used_sheet" in dir() else "", "rows_read": 0}

    if not rows:
        return {"success": True, "result": "表为空或无数据。", "error": None, "sheet_name": used_sheet, "rows_read": 0}

    def escape_md(s: str) -> str:
        return (s or "").replace("|", "\\|").replace("\n", " ")

    col_count = max(len(r) for r in rows)
    lines = []
    for i, r in enumerate(rows):
        r_padded = list(r) + [""] * (col_count - len(r))
        lines.append("| " + " | ".join(escape_md(str(c)) for c in r_padded) + " |")
    sep = "| " + " | ".join(["---"] * col_count) + " |"
    result_text = f"工作表「{used_sheet}」共 {len(rows)} 行（普适解析，未限定列）：\n\n| " + " | ".join(escape_md(str(i + 1)) for i in range(col_count)) + " |\n" + sep + "\n" + "\n".join(lines)
    return {"success": True, "result": result_text, "error": None, "sheet_name": used_sheet, "rows_read": len(rows)}


def _parse_cell_ref(ref: str) -> tuple[int, int] | None:
    """将 A1 形式转为 (row: 1-based, col: 1-based)，失败返回 None。"""
    m = re.match(r"^([A-Za-z]+)(\d+)$", (ref or "").strip())
    if not m:
        return None
    col_str, row_str = m.group(1).upper(), m.group(2)
    col = 0
    for c in col_str:
        col = col * 26 + (ord(c) - ord("A") + 1)
    try:
        row = int(row_str)
        return (row, col) if row >= 1 and col >= 1 else None
    except ValueError:
        return None


def edit_excel(
    file_path: str,
    edits: List[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """
    【普适性】编辑任意 Excel：按单元格或区域写入。不依赖报价单列结构，任意 .xlsx/.xlsm 可用。

    edits 每项为以下之一：
    - {"cell": "A1", "value": 任意}：单格写入
    - {"range": "A1:B2", "values": [[v1,v2],[v3,v4]]}：区域按行写入

    Returns:
        {"success": bool, "result": str, "error": str | None, "output_path": str}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "result": "", "error": "请安装 openpyxl: pip install openpyxl", "output_path": ""}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "result": "", "error": f"文件不存在: {path}", "output_path": ""}
    out_p = Path(output_path) if output_path else path
    if not out_p.is_absolute():
        out_p = Path(os.getcwd()) / out_p

    if not edits or not isinstance(edits, list):
        return {"success": False, "result": "", "error": "请提供 edits 数组（每项含 cell+value 或 range+values）", "output_path": ""}

    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        applied = 0
        for op in edits:
            if op.get("cell") is not None and "value" in op:
                cell_ref = str(op["cell"]).strip()
                parsed = _parse_cell_ref(cell_ref)
                if not parsed:
                    continue
                row, col = parsed
                ws.cell(row=row, column=col, value=op["value"])
                applied += 1
            elif op.get("range") is not None and op.get("values") is not None:
                range_ref = str(op["range"]).strip()
                parts = range_ref.split(":")
                if len(parts) != 2:
                    continue
                start = _parse_cell_ref(parts[0])
                end = _parse_cell_ref(parts[1])
                if not start or not end:
                    continue
                row_s, col_s = start
                row_e, col_e = end
                vals = op["values"]
                if not isinstance(vals, list):
                    continue
                for ri, row_vals in enumerate(vals):
                    if not isinstance(row_vals, list):
                        continue
                    for ci, v in enumerate(row_vals):
                        r, c = row_s + ri, col_s + ci
                        if r > row_e or c > col_e:
                            break
                        ws.cell(row=r, column=c, value=v)
                applied += 1
        wb.save(out_p)
        return {"success": True, "result": json.dumps({"applied_edits": applied, "output_path": str(out_p)}, ensure_ascii=False), "error": None, "output_path": str(out_p)}
    except Exception as e:
        return {"success": False, "result": "", "error": str(e), "output_path": ""}


def get_quote_tools_openai_format() -> list[dict]:
    """OpenAI function calling 格式：报价 Agent 工具。"""
    return [
        {
            "type": "function",
            "function": {
                "name": "extract_quotation_data",
                "description": "【报价单导向】从原始报价单 Excel 中提取报价数据。数据范围：从第2行起到「Total Excluding PPN不含税总价」所在行的上一行止，第1行为表头。返回 Markdown 表格供后续分析。调用前需确保 context 中已有 file_path（用户上传后的路径）。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "报价单 Excel 的完整路径，通常从 context.file_path 获取",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "工作表名称，不传则使用第一个/当前工作表",
                        },
                    },
                    "required": ["file_path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "fill_quotation_sheet",
                "description": "【报价单导向】将数据写入报价单 Excel 指定行。fill_items 每项含 row(Excel行号1-based)、code、quote_name、unit_price、qty、specification。写入列：G=产品编号, H=报价名称, J=规格, L=数量, N=单价。相当于 excel write_data 填表。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "报价单 Excel 路径"},
                        "fill_items": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "row": {"type": "integer", "description": "Excel 行号 1-based"},
                                    "code": {"type": "string"},
                                    "quote_name": {"type": "string"},
                                    "unit_price": {"type": "number"},
                                    "qty": {"type": "integer"},
                                    "specification": {"type": "string"},
                                },
                                "required": ["row"],
                            },
                            "description": "要回填的项列表",
                        },
                        "output_path": {"type": "string", "description": "可选，输出路径，默认覆盖原文件"},
                        "sheet_name": {"type": "string", "description": "工作表名，不传用第一个"},
                    },
                    "required": ["file_path", "fill_items"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "parse_excel_smart",
                "description": "【普适性】智能解析任意 Excel 文件：自动读取指定工作表的单元格（不限定列/行结构），适合多表头、合并单元格、不规则布局。零硬编码，返回 Markdown 表格供分析。与 MCP parse_excel_smart 语义一致。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "Excel 文件完整路径（可从 context.file_path 获取）"},
                        "sheet_name": {"type": "string", "description": "工作表名称，不传则使用第一个工作表"},
                        "max_rows": {"type": "integer", "description": "最多读取行数，默认 500", "default": 500},
                    },
                    "required": ["file_path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "edit_excel",
                "description": "【普适性】编辑任意 Excel：按单元格或区域写入，不依赖报价单列结构。edits 每项：{\"cell\": \"A1\", \"value\": 任意} 单格写入，或 {\"range\": \"A1:B2\", \"values\": [[v1,v2],[v3,v4]]} 区域按行写入。可写多格后保存到 output_path 或覆盖原文件。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "要编辑的 Excel 路径"},
                        "edits": {
                            "type": "array",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "cell": {"type": "string", "description": "单格引用，如 A1"},
                                    "value": {"description": "写入的值（字符串或数字）"},
                                    "range": {"type": "string", "description": "区域引用，如 A1:B2"},
                                    "values": {
                                        "type": "array",
                                        "items": {
                                            "type": "array",
                                            "items": {
                                                "oneOf": [
                                                    {"type": "string"},
                                                    {"type": "number"},
                                                    {"type": "boolean"},
                                                    {"type": "null"}
                                                ]
                                            }
                                        },
                                        "description": "二维数组，按行写入，每个单元格为字符串/数字/布尔或 null",
                                    },
                                },
                            },
                            "description": "编辑操作列表：单格用 cell+value，区域用 range+values",
                        },
                        "sheet_name": {"type": "string", "description": "工作表名，不传用第一个"},
                        "output_path": {"type": "string", "description": "保存路径，不传则覆盖原文件"},
                    },
                    "required": ["file_path", "edits"],
                },
            },
        },
    ]


# 凌威报价单回填列（1-based）
COL_PRODUCT_NO = 7   # G 产品编号
COL_QUOTE_NAME = 8   # H 报价名称
COL_QUOTE_SPEC = 10  # J 报价产品规格
COL_QTY_OUT = 12     # L 数量
COL_UNIT_PRICE = 14  # N 单价
COL_TOTAL = 15       # O 总价

# 4 个价格计算行（在 Total Excluding PPN 所在行及其后 3 行），金额写入列与 COL_TOTAL 一致
TOTALS_VALUE_COL = COL_TOTAL  # O 列


def fill_quotation(
    file_path: str,
    fill_items: list[dict[str, Any]],
    sheet_name: str | None = None,
    output_path: str | None = None,
    freight: float = 0.0,
) -> dict[str, Any]:
    """
    将匹配到的产品信息回填到报价单 Excel。
    未匹配项写「无货」；写完后更新底部 4 个价格计算行。

    Args:
        file_path: 原始报价单路径
        fill_items: 每项含 row, code, quote_name, unit_price, qty, specification；code="无货" 表示未匹配
        sheet_name: 工作表名，默认第一个
        output_path: 输出路径，默认覆盖原文件（建议调用方传副本路径）
        freight: 运费，默认 0

    Returns:
        {"success": bool, "output_path": str, "filled_count": int, "error": str | None}
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "output_path": "", "filled_count": 0, "error": "请安装 openpyxl"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists():
        return {"success": False, "output_path": "", "filled_count": 0, "error": f"文件不存在: {path}"}
    out_p = Path(output_path) if output_path else path
    if not out_p.is_absolute():
        out_p = Path(os.getcwd()) / out_p
    try:
        wb = openpyxl.load_workbook(path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]
        filled = 0
        total_excluding_ppn = 0.0
        for it in fill_items:
            row_num = it.get("row")
            if not row_num:
                continue
            code = it.get("code")
            if code:
                ws.cell(row=row_num, column=COL_PRODUCT_NO, value=str(code))
                filled += 1
            if it.get("quote_name"):
                ws.cell(row=row_num, column=COL_QUOTE_NAME, value=str(it["quote_name"]))
            if it.get("unit_price") is not None:
                ws.cell(row=row_num, column=COL_UNIT_PRICE, value=float(it["unit_price"]))
            if it.get("qty") is not None:
                ws.cell(row=row_num, column=COL_QTY_OUT, value=int(it["qty"]))
            if it.get("specification"):
                ws.cell(row=row_num, column=COL_QUOTE_SPEC, value=str(it["specification"]))
            up = it.get("unit_price")
            q = it.get("qty")
            if up is not None and q is not None and code and str(code) != "无货":
                row_total = float(up) * int(q)
                ws.cell(row=row_num, column=COL_TOTAL, value=round(row_total, 2))
                total_excluding_ppn += row_total
            elif up is not None and q is not None and (not code or str(code) == "无货"):
                ws.cell(row=row_num, column=COL_TOTAL, value=0)
        ppn = round(total_excluding_ppn * 0.11, 2)
        total_including = round(total_excluding_ppn + ppn + float(freight), 2)
        total_row_1based = None
        for row in ws.iter_rows():
            row_idx = row[0].row if row else 0
            for cell in row:
                if TOTAL_ROW_MARKER in _cell_value(cell):
                    total_row_1based = row_idx
                    break
            if total_row_1based is not None:
                break
        if total_row_1based is not None:
            ws.cell(row=total_row_1based, column=TOTALS_VALUE_COL, value=round(total_excluding_ppn, 2))
            ws.cell(row=total_row_1based + 1, column=TOTALS_VALUE_COL, value=ppn)
            ws.cell(row=total_row_1based + 2, column=TOTALS_VALUE_COL, value=float(freight))
            ws.cell(row=total_row_1based + 3, column=TOTALS_VALUE_COL, value=total_including)
        wb.save(out_p)
        return {"success": True, "output_path": str(out_p), "filled_count": filled, "error": None}
    except Exception as e:
        return {"success": False, "output_path": "", "filled_count": 0, "error": str(e)}


def execute_quote_tool(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    """执行报价 Agent 工具，返回 {success, result, error, ...} 或 {success, items, error, rows_count}。"""
    if name == "extract_quotation_data":
        file_path = (arguments.get("file_path") or "").strip()
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        if not file_path:
            return {"success": False, "result": "", "error": "请提供 file_path", "rows_count": 0}
        return extract_quotation_data(file_path=file_path, sheet_name=sheet_name)
    if name == "fill_quotation_sheet":
        file_path = (arguments.get("file_path") or "").strip()
        fill_items = arguments.get("fill_items") or []
        output_path = (arguments.get("output_path") or "").strip() or None
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        if not file_path:
            return {"success": False, "result": "", "error": "请提供 file_path"}
        if not fill_items or not isinstance(fill_items, list):
            return {"success": False, "result": "", "error": "请提供 fill_items 数组"}
        out = fill_quotation(
            file_path=file_path,
            fill_items=fill_items,
            sheet_name=sheet_name,
            output_path=output_path,
        )
        if out.get("success"):
            return {"success": True, "result": json.dumps({"filled_count": out["filled_count"], "output_path": out["output_path"]}, ensure_ascii=False), "error": None}
        return {"success": False, "result": "", "error": out.get("error", "填表失败")}
    if name == "parse_excel_smart":
        fp = (arguments.get("file_path") or "").strip()
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        max_rows = arguments.get("max_rows")
        if max_rows is None:
            max_rows = 500
        try:
            max_rows = int(max_rows)
        except (TypeError, ValueError):
            max_rows = 500
        if not fp:
            return {"success": False, "result": "", "error": "请提供 file_path", "rows_read": 0}
        out = parse_excel_smart(file_path=fp, sheet_name=sheet_name, max_rows=max_rows)
        if out.get("success"):
            return {"success": True, "result": out["result"], "error": None, "rows_read": out.get("rows_read", 0)}
        return {"success": False, "result": "", "error": out.get("error", "解析失败"), "rows_read": 0}
    if name == "edit_excel":
        fp = (arguments.get("file_path") or "").strip()
        edits = arguments.get("edits") or []
        sheet_name = (arguments.get("sheet_name") or "").strip() or None
        output_path = (arguments.get("output_path") or "").strip() or None
        if not fp:
            return {"success": False, "result": "", "error": "请提供 file_path", "output_path": ""}
        out = edit_excel(file_path=fp, edits=edits, sheet_name=sheet_name, output_path=output_path)
        if out.get("success"):
            return {"success": True, "result": out["result"], "error": None, "output_path": out.get("output_path", "")}
        return {"success": False, "result": "", "error": out.get("error", "编辑失败"), "output_path": ""}
    return {"success": False, "result": "", "error": f"未知工具: {name}", "rows_count": 0}
