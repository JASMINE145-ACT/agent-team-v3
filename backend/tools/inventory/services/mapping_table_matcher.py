"""
询价→产品映射表匹配：基于「字段名 + 规格」的模糊查询，逻辑与万鼎一致但数据源为映射表 Excel。
若 query 命中映射表则取 top3 后由 LLM 选 1 个；未命中则走万鼎价格库。
映射表列：A=询价货物名称 B=询价规格型号 C=产品编号 D=报价名称；B 可为空（规格合在 A 中）。
"""

from __future__ import annotations

import logging
import re
import threading
from pathlib import Path
from typing import Any, List, Optional, Tuple

import pandas as pd

from backend.tools.inventory.services.wanding_fuzzy_matcher import (
    _SINGLE_CHAR_WEIGHT,
    _SORTED_SYNONYMS,
    _SYNONYM_TO_GROUP,
    _expand_keyword_with_synonyms,
    _expand_token_with_synonyms_and_units,
    _normalize,
    _normalize_keyword_terms,
    _split_tokens,
)

logger = logging.getLogger(__name__)

_MAPPING_DF_CACHE: Optional[pd.DataFrame] = None
_MAPPING_DF_PATH: Optional[str] = None
_MAPPING_CACHE_LOCK = threading.Lock()


def load_mapping_df(path: str | Path) -> pd.DataFrame:
    """
    加载映射表 Excel，Sheet1，第 1 行为表头，数据从第 2 行起。
    列 A=询价货物名称 B=询价规格型号 C=产品编号 D=报价名称。
    返回 DataFrame：search_text（A+B 合并，用于匹配）, code, matched_name，
    以及预计算的 norm_text / spec_tokens 列。
    """
    global _MAPPING_DF_CACHE, _MAPPING_DF_PATH
    path_str = str(path)
    if _MAPPING_DF_CACHE is not None and _MAPPING_DF_PATH == path_str:
        return _MAPPING_DF_CACHE

    with _MAPPING_CACHE_LOCK:
        # double-checked locking
        if _MAPPING_DF_CACHE is not None and _MAPPING_DF_PATH == path_str:
            return _MAPPING_DF_CACHE

        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl 未安装，映射表匹配不可用")
            return pd.DataFrame()

        p = Path(path)
        if not p.is_absolute():
            root = Path(__file__).resolve().parent.parent.parent.parent.parent
            p = root / p
        if not p.exists():
            logger.warning("映射表不存在: %s", p)
            return pd.DataFrame()

        rows: List[dict] = []
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb.active or wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(max_col=4, min_row=2)):
                cells = [getattr(c, "value", None) for c in row]
                if len(cells) < 4:
                    continue
                field_name = str(cells[0] or "").strip()
                spec = str(cells[1] or "").strip() if cells[1] is not None else ""
                code = cells[2]
                if code is None:
                    continue
                code = str(code).strip()
                matched_name = str(cells[3] or "").strip()
                if not field_name and not code:
                    continue
                search_text = f"{field_name} {spec}".strip() if spec else field_name
                rows.append({"search_text": search_text, "code": code, "matched_name": matched_name})
            wb.close()
        except Exception as e:
            logger.warning("加载映射表失败: %s", e)
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        if not df.empty:
            # 预计算 normalized text 和规格 token 集，避免 search_mapping_fuzzy 每行重算
            df["norm_text"] = df["search_text"].apply(_normalize)
            df["spec_tokens"] = df["search_text"].apply(
                lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
            )
            _MAPPING_DF_CACHE = df
            _MAPPING_DF_PATH = path_str
        return df


def search_mapping_fuzzy(
    df: pd.DataFrame,
    keyword: str,
) -> List[Tuple[dict[str, Any], float]]:
    """
    与万鼎相同的模糊逻辑：token + 同义词 + 规格等价，对 df 的 search_text 列打分。
    返回 [(row_dict, score), ...] 按 score 降序，row_dict = {code, matched_name, unit_price: 0}。

    优化点（与 search_fuzzy 一致）：
    - 使用 load_mapping_df 预计算的 norm_text / spec_tokens 列
    - q_eq 提出行循环，set 交集替代 any()
    - 单字 token 降权（_SINGLE_CHAR_WEIGHT）
    """
    if df.empty or "search_text" not in df.columns:
        return []

    keyword = _normalize_keyword_terms((keyword or "").strip())
    if not keyword:
        return []

    has_precomputed = "norm_text" in df.columns and "spec_tokens" in df.columns
    results: dict[str, Tuple[dict, float]] = {}

    for kw in _expand_keyword_with_synonyms(keyword):
        norm_kw = _normalize(kw)
        chinese_tokens = _split_tokens(norm_kw)
        material_tokens = re.findall(r"pvc|ppr|pe|hdpe", norm_kw)
        query_size_tokens = {t for t in chinese_tokens if re.search(r"\d", t) and not t.endswith("°")}
        query_text_tokens = {
            t for t in chinese_tokens if not (re.search(r"\d", t) and not t.endswith("°"))
        }
        query_material = material_tokens[0] if material_tokens else None

        # q_eq 提出行循环
        spec_equivs: dict[str, frozenset] = {
            q_spec: _expand_token_with_synonyms_and_units(q_spec, material=query_material)
            for q_spec in query_size_tokens
        }

        multi_text = {t for t in query_text_tokens if len(t) > 1}
        single_text = {t for t in query_text_tokens if len(t) == 1}
        total_weight = (
            len(query_size_tokens)
            + len(multi_text)
            + len(single_text) * _SINGLE_CHAR_WEIGHT
        )

        for row in df.itertuples(index=False):
            raw_text = str(getattr(row, "search_text", ""))
            if not raw_text:
                continue
            code = str(getattr(row, "code", "")).strip()
            matched_name = str(getattr(row, "matched_name", "")).strip()

            if has_precomputed:
                normalized_text: str = row.norm_text
                product_specs: frozenset = row.spec_tokens
            else:
                normalized_text = _normalize(raw_text)
                product_specs = frozenset(
                    t for t in _split_tokens(raw_text) if re.search(r"\d", t)
                )

            if material_tokens and not all(m.lower() in normalized_text for m in material_tokens):
                continue

            size_hits = sum(1 for q_eq in spec_equivs.values() if q_eq & product_specs)
            if query_size_tokens and size_hits == 0:
                continue

            def _text_match(t: str) -> bool:
                return t.lower() in normalized_text or (t == "度" and "°" in normalized_text)

            multi_hits = sum(1 for t in multi_text if _text_match(t))
            single_hits = sum(1 for t in single_text if _text_match(t))
            if query_text_tokens and (multi_hits + single_hits) == 0:
                continue

            hit_weight = size_hits + multi_hits + single_hits * _SINGLE_CHAR_WEIGHT
            score = hit_weight / total_weight if total_weight > 0 else 0.0

            if score > 0 and (code not in results or score > results[code][1]):
                results[code] = (
                    {"code": code, "matched_name": matched_name, "unit_price": 0.0},
                    score,
                )

    out = list(results.values())
    out.sort(key=lambda x: x[1], reverse=True)
    return out


def match_mapping_top_candidates(
    keywords: str,
    mapping_path: Optional[str] = None,
    top_k: int = 3,
) -> List[dict[str, Any]]:
    """
    映射表模糊匹配，返回前 top_k 个候选（每项含 code, matched_name, unit_price=0）。
    无命中返回空列表。
    """
    if mapping_path is None:
        try:
            from backend.tools.inventory.config import config
            mapping_path = getattr(config, "MAPPING_TABLE_PATH", None)
        except Exception:
            return []
    if not mapping_path:
        return []

    df = load_mapping_df(mapping_path)
    if df.empty:
        return []

    pairs = search_mapping_fuzzy(df, keywords)[:top_k]
    return [row_dict for row_dict, _ in pairs]
