"""
万鼎价格库匹配 - DataBase-style 模糊逻辑

仅此一种查询逻辑：token + 同义词扩展 + 规格等价 + score 排序。
借鉴 DataBase- 项目 search_with_keywords(strict=False, return_score=True)。
"""

from __future__ import annotations

import logging
import re
import threading
from pathlib import Path
from typing import Any, List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

PRICE_COLS = {"A": 8, "B": 10, "C": 12, "D": 14}

SYNONYM_GROUPS = [
    {"直接", "直接头", "直通", "直通接头"},
    {"变径", "异径"},
    {"大小头", "异径直通", "异径套", "变径直接", "异径直接"},
    {"内丝", "内螺纹"}, {"外丝", "外螺纹"},
    {"锁母", "锁扣", "管接头"},
    {"止回阀", "截止阀"},
    {"穿线管", "电线管"},
    {"半弯", "弯头"},
    {"承插", "承插式"},
    {"堵头", "管帽"},
]

# 模块级预计算，避免每次调用时重建（SYNONYM_GROUPS 不变时永远有效）
_SYNONYM_TO_GROUP: dict[str, frozenset] = {
    syn: frozenset(group) for group in SYNONYM_GROUPS for syn in group
}
_SORTED_SYNONYMS: list[str] = sorted(_SYNONYM_TO_GROUP.keys(), key=len, reverse=True)

# 单字 token（如「三」「通」）在打分中的权重，相对于多字 token 的 1.0
_SINGLE_CHAR_WEIGHT = 0.5

# 询价关键词中的英文/印尼语/口语 → 中文品名，用于筛选时命中库内品名（与 wanding_business_knowledge.md 保持一致，便于 LLM 选型有思路）
QUERY_TERM_TO_CHINESE = [
    ("4 cabang", "管四通圆接线盒"),
    ("conduit", "电线管"), ("counduit", "电线管"), ("pipa", "管"),
    ("socket", "管直通"), ("套筒", "管直通"),
    ("klem", "管夹"),
    ("cabang", "四通"), ("tdust", "四通"),
    ("热熔器", "焊接机"), ("热熔机", "焊接机"), ("熔接器", "焊接机"),
    ("四通接线盒", "管四通圆接线盒"),
    ("马鞍卡", "管夹"),
]


def _normalize_keyword_terms(keywords: str) -> str:
    """将询价中的英文/印尼语替换为中文品名词，便于筛选命中库内品名。"""
    s = (keywords or "").strip()
    for eng, ch in QUERY_TERM_TO_CHINESE:
        s = re.sub(r"\b" + re.escape(eng) + r"\b", ch, s, flags=re.I)
    return s.strip()


# 业务知识中【字段匹配同义与规格】规则缓存，供字段匹配阶段使用（与 LLM 选型共用同一 knowledge 文件）
_FIELD_MATCHING_RULES_CACHE: dict = {}  # {"path": str, "mtime": float|None, "rules": [(sources, targets), ...]}


def _load_field_matching_rules_from_knowledge() -> List[tuple[List[str], List[str]]]:
    """
    从 wanding_business_knowledge.md 的【字段匹配同义与规格】段落解析规则，
    用于字段匹配阶段同义扩展，提高命中率。返回 [(source_terms, target_terms), ...]。
    """
    global _FIELD_MATCHING_RULES_CACHE
    try:
        from backend.tools.inventory.config import config
        path_str = getattr(config, "WANDING_BUSINESS_KNOWLEDGE_PATH", None)
        if not path_str:
            return []
        p = Path(path_str)
        if not p.exists():
            return []
        mtime: Optional[float] = None
        try:
            mtime = p.stat().st_mtime
        except OSError:
            pass
        if (
            _FIELD_MATCHING_RULES_CACHE.get("path") == path_str
            and _FIELD_MATCHING_RULES_CACHE.get("mtime") == mtime
        ):
            return _FIELD_MATCHING_RULES_CACHE.get("rules") or []
        content = p.read_text(encoding="utf-8")
        rules: List[tuple[List[str], List[str]]] = []
        in_section = False
        for line in content.splitlines():
            line = line.strip()
            if "【字段匹配" in line or "【字段匹配同义" in line:
                in_section = True
                continue
            if in_section and line.startswith("【"):
                break
            if not in_section:
                continue
            # 解析 "- 源词 源词 → 检索词 检索词" 或 "  - ... → ..."
            if line.startswith("-"):
                line = line.lstrip("-").strip()
            if "→" in line:
                left, _, right = line.partition("→")
            elif "->" in line:
                left, _, right = line.partition("->")
            else:
                continue
            sources = [t.strip() for t in left.split() if t.strip()]
            targets = [t.strip() for t in right.split() if t.strip()]
            if sources and targets:
                rules.append((sources, targets))
        _FIELD_MATCHING_RULES_CACHE = {"path": path_str, "mtime": mtime, "rules": rules}
        return rules
    except Exception as e:
        logger.debug("加载字段匹配规则失败: %s", e)
        return []


def _apply_knowledge_expansion(keywords: str) -> str:
    """
    根据业务知识【字段匹配同义与规格】规则，在字段匹配前扩展询价词，
    使口语/同义词能命中库内品名（如 直接→直通 排水、热熔器→焊接机）。
    """
    if not (keywords or "").strip():
        return keywords
    rules = _load_field_matching_rules_from_knowledge()
    added: List[str] = []
    kw_lower = (keywords or "").lower()
    for sources, targets in rules:
        for src in sources:
            if src.lower() in kw_lower or re.search(re.escape(src), keywords, re.I):
                added.extend(targets)
                break
    if not added:
        return keywords.strip()
    return (keywords.strip() + " " + " ".join(added)).strip()

MM_TO_INCH = {
    "16": '1/2"', "20": '3/4"', "25": '1"', "32": '1-1/4"', "40": '1-1/2"',
    "50": '2"', "65": '2-1/2"', "75": '3"', "100": '4"', "125": '5"',
    "150": '6"', "200": '8"', "250": '10"', "300": '12"',
}
INCH_TO_MM = {v: k for k, v in MM_TO_INCH.items()}


def _normalize(s: str) -> str:
    s = (s or "").lower().strip()
    s = s.replace("－", "-").replace("—", "-").replace("（", "(").replace("）", ")")
    s = re.sub(r"[_\t]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _get_synonym_words(word: str) -> frozenset:
    return _SYNONYM_TO_GROUP.get(word, frozenset({word}))


def _expand_unit_tokens(token: str, material: Optional[str] = None) -> set:
    eqs = {token}
    if token.startswith("dn"):
        num = token[2:]
        if num in MM_TO_INCH:
            eqs.add(MM_TO_INCH[num])
        eqs.add(num)
        return eqs
    if token.isdigit() and token in MM_TO_INCH:
        eqs.add("dn" + token)
        eqs.add(MM_TO_INCH[token])
        return eqs
    if token in INCH_TO_MM:
        eqs.add(INCH_TO_MM[token])
        eqs.add("dn" + INCH_TO_MM[token])
        return eqs
    return eqs


def _expand_token_with_synonyms_and_units(token: str, material: Optional[str] = None) -> set:
    synonyms = _get_synonym_words(token)
    expanded: set = set()
    for syn in synonyms:
        expanded |= _expand_unit_tokens(syn, material=material)
    return expanded


def _split_tokens(text: str) -> List[str]:
    text = _normalize(text)
    tokens: List[str] = []
    for m in re.finditer(r"dn\s*(\d+)", text, re.I):
        tokens.append("dn" + m.group(1))
        tokens.append(m.group(1))
    text = re.sub(r"dn\s*\d+", " ", text, flags=re.I)
    for m in re.finditer(r"\d+(?:\.\d+)?", text):
        tokens.append(m.group())
    text = re.sub(r"\d+(?:\.\d+)?", " ", text)
    for m in re.finditer(r"[\u4e00-\u9fff]+", text):
        tok = m.group()
        tokens.append(tok)
        # 长中文整段（如「度弯头带检查口」）在品名中常被括号/符号隔开，拆成单字以便「度」→「°」、「弯头」「带」「检查口」等能分别命中
        if len(tok) > 2:
            for c in tok:
                tokens.append(c)
    return list(dict.fromkeys(tokens))


def _expand_keyword_with_synonyms(keyword: str) -> List[str]:
    queries: set = {keyword}
    for syn in _SORTED_SYNONYMS:
        new_queries: set = set()
        for q in queries:
            if syn in q:
                for replacement in _SYNONYM_TO_GROUP[syn]:
                    new_queries.add(q.replace(syn, replacement))
        if new_queries:
            queries.update(new_queries)
    return list(queries)


def search_fuzzy(
    df: pd.DataFrame,
    keyword: str,
    field: str = "Describrition",
) -> List[tuple[dict[str, Any], float]]:
    """
    DataBase-style fuzzy search.
    Returns [(row_dict, score), ...] sorted by score desc.
    row_dict: {code, matched_name, unit_price}

    优化点：
    - 使用 load_wanding_df 预计算的 norm_text / spec_tokens 列，消除每行的 regex 开销
    - q_eq 提出行循环，每次查询只计算一次
    - set 交集（q_eq & product_specs）替代 any(eq in set for eq in set)
    - 单字 token 权重 _SINGLE_CHAR_WEIGHT（0.5），避免单字命中过度拉高得分
    """
    results: dict = {}
    has_precomputed = "norm_text" in df.columns and "spec_tokens" in df.columns

    for kw in _expand_keyword_with_synonyms(keyword.strip()):
        norm_kw = _normalize(kw)
        chinese_tokens = _split_tokens(norm_kw)
        material_tokens = re.findall(r"pvc|ppr|pe|hdpe", norm_kw)
        query_size_tokens = {t for t in chinese_tokens if re.search(r"\d", t) and not t.endswith("°")}
        query_text_tokens = {
            t for t in chinese_tokens if not (re.search(r"\d", t) and not t.endswith("°"))
        }
        query_material = material_tokens[0] if material_tokens else None

        # q_eq 提出行循环：只依赖 q_spec + query_material，与当前行无关
        spec_equivs: dict[str, frozenset] = {
            q_spec: _expand_token_with_synonyms_and_units(q_spec, material=query_material)
            for q_spec in query_size_tokens
        }

        # 按单字/多字分类，单字在分母中按 _SINGLE_CHAR_WEIGHT 计入
        multi_text = {t for t in query_text_tokens if len(t) > 1}
        single_text = {t for t in query_text_tokens if len(t) == 1}
        total_weight = (
            len(query_size_tokens)
            + len(multi_text)
            + len(single_text) * _SINGLE_CHAR_WEIGHT
        )

        for row in df.itertuples(index=False):
            row_id = getattr(row, "Material", getattr(row, "Describrition", str(row)))
            raw_text = str(getattr(row, field, ""))

            # 使用预计算列，fallback 到实时计算（兼容未预计算的 df）
            if has_precomputed:
                normalized_text: str = row.norm_text
                product_specs: frozenset = row.spec_tokens
            else:
                normalized_text = _normalize(raw_text)
                product_specs = frozenset(
                    t for t in _split_tokens(raw_text) if re.search(r"\d", t)
                )

            if material_tokens and not all(m.lower() in normalized_text for m in material_tokens):
                continue

            # set 交集替代 any(eq in product_specs for eq in q_eq)
            size_hits = sum(1 for q_eq in spec_equivs.values() if q_eq & product_specs)
            if query_size_tokens and size_hits == 0:
                continue

            # 多字命中（权重 1.0）+ 单字命中（权重 _SINGLE_CHAR_WEIGHT）
            def _text_match(t: str) -> bool:
                return t.lower() in normalized_text or (t == "度" and "°" in normalized_text)

            multi_hits = sum(1 for t in multi_text if _text_match(t))
            single_hits = sum(1 for t in single_text if _text_match(t))
            # 过滤用原始命中数（单字也算），得分用加权值
            if query_text_tokens and (multi_hits + single_hits) == 0:
                continue

            hit_weight = size_hits + multi_hits + single_hits * _SINGLE_CHAR_WEIGHT
            score = hit_weight / total_weight if total_weight > 0 else 0.0

            if score > 0 and (row_id not in results or score > results[row_id][1]):
                row_dict: dict[str, Any] = {
                    "code": str(getattr(row, "Material", "")),
                    "matched_name": raw_text,
                }
                if hasattr(row, "unit_price"):
                    row_dict["unit_price"] = getattr(row, "unit_price", 0.0)
                results[row_id] = (row_dict, score)

    out = list(results.values())
    out.sort(key=lambda x: x[1], reverse=True)
    return out


def _load_one_sheet(ws, price_col: int) -> list[dict]:
    """从已打开的 worksheet 读出一张表的行（Material, Describrition, unit_price）。"""
    rows = []
    for row in ws.iter_rows(max_col=16):
        cells = [getattr(c, "value", None) for c in row]
        if len(cells) > 2 and cells[2]:
            up = 0.0
            if len(cells) > price_col and cells[price_col] is not None:
                try:
                    up = float(cells[price_col])
                except (ValueError, TypeError):
                    pass
            rows.append({
                "Material": str(cells[1] or "").strip(),
                "Describrition": str(cells[2] or "").strip(),
                "unit_price": up,
            })
    return rows


def load_wanding_df(
    path: str | Path,
    sheet_name: str = "管材",
    customer_level: str = "B",
) -> pd.DataFrame:
    """Load 万鼎 price library as DataFrame. 默认加载「管材」+「国标管件」两个 sheet 并合并，以便匹配带检查口弯头、管帽等国标管件。"""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl 未安装，万鼎模糊匹配不可用")
        return pd.DataFrame()

    p = Path(path)
    if p.is_absolute() and p.exists():
        pass
    elif not p.is_absolute():
        root = Path(__file__).resolve().parent.parent.parent
        p = root / p
    if not p.exists():
        logger.warning("万鼎价格库不存在: %s", p)
        return pd.DataFrame()

    level = (customer_level or "B").strip().upper()
    price_col = PRICE_COLS.get(level, PRICE_COLS["B"])

    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        all_rows: list[dict] = []
        # 先加载管材
        ws_guan = wb["管材"] if "管材" in wb.sheetnames else (wb.active or wb[wb.sheetnames[0]])
        all_rows.extend(_load_one_sheet(ws_guan, price_col))
        # 若存在国标管件 sheet，一并加载（8020020643 带检查口弯头、8020020205 管帽等在此表）
        if "国标管件" in wb.sheetnames:
            all_rows.extend(_load_one_sheet(wb["国标管件"], price_col))
        wb.close()
        df = pd.DataFrame(all_rows)
        if not df.empty:
            # 预计算 normalized text 和规格 token 集，避免 search_fuzzy 每行重算
            df["norm_text"] = df["Describrition"].apply(_normalize)
            df["spec_tokens"] = df["Describrition"].apply(
                lambda t: frozenset(tok for tok in _split_tokens(t) if re.search(r"\d", tok))
            )
        return df
    except Exception as e:
        logger.warning("加载万鼎价格库失败: %s", e)
        return pd.DataFrame()


# 缓存 DataFrame，按 path:level 隔离
_df_cache: dict[str, pd.DataFrame] = {}
_df_cache_lock = threading.Lock()


def _get_cached_df(path, customer_level: str) -> pd.DataFrame:
    """线程安全地获取（或加载）指定 path+level 的 DataFrame。"""
    level = (customer_level or "B").strip().upper()
    cache_key = f"{path}:{level}"
    if cache_key in _df_cache:
        return _df_cache[cache_key]
    with _df_cache_lock:
        if cache_key not in _df_cache:
            _df_cache[cache_key] = load_wanding_df(path, customer_level=level)
    return _df_cache[cache_key]


def match_fuzzy(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
) -> Optional[dict[str, Any]]:
    """
    DataBase-style 模糊匹配，返回最佳单结果。
    返回 {code, matched_name, unit_price} 或 None。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换（SYNONYM_GROUPS、QUERY_TERM_TO_CHINESE）。
    """
    keywords = (keywords or "").strip()
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _normalize_keyword_terms(keywords)
    if not keywords:
        return None

    from backend.tools.inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty:
        return None

    results = search_fuzzy(df, keywords)
    if not results:
        return None

    row_dict, _ = results[0]
    return {
        "code": (row_dict.get("code") or "").strip(),
        "matched_name": (row_dict.get("matched_name") or "")[:200],
        "unit_price": float(row_dict.get("unit_price", 0) or 0),
    }


def match_fuzzy_candidates(
    keywords: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
    max_candidates: int = 20,
    max_score_tiers: Optional[int] = None,
) -> List[dict[str, Any]]:
    """
    返回候选列表，每项含 code, matched_name, unit_price, score。
    - max_score_tiers 为 None：按分数排序取前 max_candidates 条。
    - max_score_tiers 为 N（如 2）：取分数前 N 档，每档全部返回（如 top1 有 3 条、top2 有 2 条则共 5 条）。
    先按业务知识【字段匹配补充规则】扩展检索词，再做同义/外语替换。
    """
    keywords = (keywords or "").strip()
    keywords = _apply_knowledge_expansion(keywords)
    keywords = _normalize_keyword_terms(keywords)
    if not keywords:
        return []

    from backend.tools.inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty:
        return []

    results = search_fuzzy(df, keywords)
    if max_score_tiers is not None and max_score_tiers > 0:
        # 取前 max_score_tiers 个分数档，每档全部返回
        tiers: List[float] = []
        for _rd, score in results:
            if score not in tiers:
                tiers.append(score)
                if len(tiers) >= max_score_tiers:
                    break
        results = [(rd, s) for rd, s in results if s in tiers]
    else:
        results = results[:max_candidates]
    out = []
    for row_dict, score in results:
        out.append({
            "code": (row_dict.get("code") or "").strip(),
            "matched_name": (row_dict.get("matched_name") or "")[:200],
            "unit_price": float(row_dict.get("unit_price", 0) or 0),
            "score": round(score, 4),
        })
    return out


def get_wanding_price_by_code(
    code: str,
    customer_level: str = "B",
    price_library_path: Optional[str | Path] = None,
) -> Optional[dict[str, Any]]:
    """
    按产品编号（Material）在万鼎价格表中精确查找，返回该 code 的单价及名称。
    用于历史匹配拿到 code 后，从万鼎表把价格补全。
    返回 {code, matched_name, unit_price} 或 None（万鼎表无此 code）。
    """
    code = (code or "").strip()
    if not code:
        return None
    from backend.tools.inventory.config import config
    path = price_library_path or config.PRICE_LIBRARY_PATH
    df = _get_cached_df(path, customer_level)
    if df.empty or "Material" not in df.columns:
        return None
    # Material 列为产品编号
    row = df[df["Material"].astype(str).str.strip() == code]
    if row.empty:
        return None
    r = row.iloc[0]
    return {
        "code": code,
        "matched_name": str(r.get("Describrition", "") or "")[:200],
        "unit_price": float(r.get("unit_price", 0) or 0),
    }
