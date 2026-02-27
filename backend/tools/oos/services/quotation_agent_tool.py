# 报价单 Agent 工具：按格解析 Excel、统计无货；抓取无货记录供 LLM 观察；按 LLM 选择持久化到 DB
# 遵循 opencode_style_agent/quotation_tools.py：按单元格迭代、无货关键词匹配、返回 count/rows_by_sheet/message
# 与 inventory_agent_tools 对齐：统一 execute_*_tool 入口，带超时封装
from __future__ import annotations

import logging
import os
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from pathlib import Path
from typing import Any

from backend.tools.oos.config import OUT_OF_STOCK_KEYWORDS

logger = logging.getLogger(__name__)


def _cell_value(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def analyze_quotation_excel(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    按单元格解析 Excel，统计包含「无货」等关键词的单元格数量及行号。
    与 opencode_style_agent/quotation_tools 逻辑一致。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "请安装 openpyxl: pip install openpyxl"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    # 与 opencode_style_agent 一致：模型可能只输出 .xlsx，实际文件名为 xxx.xlsx...xlsx
    if not path.exists() and not file_path.endswith(".xlsx...xlsx"):
        alt = Path(str(path) + "...xlsx")
        if alt.exists():
            path = alt
    if not path.exists():
        return {"success": False, "error": f"文件不存在: {path}"}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "error": "仅支持 .xlsx / .xlsm"}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"打开 Excel 失败: {e}"}

    try:
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        total_count = 0
        rows_by_sheet: dict[str, list[int]] = {}
        detail_lines: list[str] = []

        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            rows_with_wuhuo: list[int] = []
            for row in ws.iter_rows():
                row_idx = row[0].row if row else 0
                for cell in row:
                    val = _cell_value(cell)
                    if not val:
                        continue
                    val_lower = val.lower()
                    if any(kw.lower() in val_lower for kw in OUT_OF_STOCK_KEYWORDS):
                        rows_with_wuhuo.append(row_idx)
                        break
            rows_with_wuhuo = sorted(set(rows_with_wuhuo))
            count = len(rows_with_wuhuo)
            total_count += count
            rows_by_sheet[sn] = rows_with_wuhuo
            if count > 0:
                detail_lines.append(
                    f"Sheet「{sn}」: {count} 条无货，行号: {rows_with_wuhuo[:30]}{'...' if len(rows_with_wuhuo) > 30 else ''}"
                )

        wb.close()
        message = f"共 {total_count} 条无货。"
        if detail_lines:
            message += " " + "; ".join(detail_lines)

        return {
            "success": True,
            "result": {
                "count": total_count,
                "rows_by_sheet": rows_by_sheet,
                "message": message,
            },
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def get_out_of_stock_records(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    只抓取无货行的原始行数据（不做列映射）。返回每行的 cells 列表及表头行，由 LLM 自行做列映射并决定要存库的记录。
    """
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "请安装 openpyxl: pip install openpyxl"}

    path = Path(file_path)
    if not path.is_absolute():
        path = Path(os.getcwd()) / path
    if not path.exists() and not file_path.endswith(".xlsx...xlsx"):
        alt = Path(str(path) + "...xlsx")
        if alt.exists():
            path = alt
    if not path.exists():
        return {"success": False, "error": f"文件不存在: {path}"}
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"success": False, "error": "仅支持 .xlsx / .xlsm"}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"打开 Excel 失败: {e}"}

    records: list[dict[str, Any]] = []
    header_lines: list[str] = []
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    global_index = 0

    try:
        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            # 表头：取首行原始 cells 供 LLM 参考列含义（不固化映射）
            first_row = ws[1]
            header_cells = [_cell_value(c) for c in first_row]
            header_lines.append(f"Sheet「{sn}」首行(表头): " + " | ".join(str(v) for v in header_cells if v))

            for row in ws.iter_rows():
                row_idx = row[0].row if row else 0
                if row_idx < 1:
                    continue
                cells = [_cell_value(c) for c in row]
                row_text = " ".join(cells).lower()
                if any(kw.lower() in row_text for kw in OUT_OF_STOCK_KEYWORDS):
                    rec = {
                        "index": global_index,
                        "sheet_name": sn,
                        "row": row_idx,
                        "cells": cells,
                    }
                    records.append(rec)
                    global_index += 1
        wb.close()

        lines = [
            "共 %d 条无货行，以下为原始行数据（不做列映射）。请你根据表头自行判断每列含义，并决定哪些是有效产品无货记录、每条对应的 product_name/specification/unit/quantity，再调用 persist_out_of_stock_records(file_name, records) 传入你要存库的列表，records 为你映射后的 [{product_name, specification, unit, quantity}, ...]。"
            % len(records),
            "",
            *header_lines,
            "",
        ]
        for r in records[:100]:
            cells_preview = " | ".join(str(x)[:30] for x in (r["cells"] or [])[:15])
            if len(r.get("cells") or []) > 15:
                cells_preview += " ..."
            lines.append(f"  index={r['index']} sheet={r['sheet_name']} row={r['row']}: {cells_preview}")
        if len(records) > 100:
            lines.append(f"  ... 仅显示前 100 条，共 {len(records)} 条。")
        message = "\n".join(lines)
        return {
            "success": True,
            "result": {
                "records": records,
                "count": len(records),
                "message": message,
            },
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def persist_out_of_stock_records(
    file_name: str,
    records: list[dict[str, Any]],
    sheet_name: str = "",
) -> dict[str, Any]:
    """
    将无货记录写入持久化数据库，落库后展示在「无货产品列表」表。
    每条 record 格式与列表一致：product_name（必填）, specification, unit, quantity；
    可选 sheet_name（单条覆盖时使用）；未传则用顶层 sheet_name。
    表内 id、uploaded_at、product_key、count 由系统自动生成。
    """
    from backend.tools.oos.models.models import OutOfStockProduct
    from backend.tools.oos.services.oos_repository import insert_oos_record
    from .data_service import DataService

    data_service = DataService()
    inserted = 0
    errors: list[str] = []
    default_sheet = (sheet_name or "").strip()
    for r in records:
        if not isinstance(r, dict):
            continue
        product_name = (r.get("product_name") or "").strip()
        if not product_name:
            continue
        try:
            qty = float(r.get("quantity") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        row_sheet = (r.get("sheet_name") or default_sheet or "").strip() if isinstance(r.get("sheet_name"), str) else default_sheet
        product = OutOfStockProduct(
            product_name=product_name,
            specification=(r.get("specification") or "").strip() or None,
            unit=(r.get("unit") or "").strip() or "",
            quantity=qty,
        )
        try:
            count = data_service.insert_record(product, row_sheet, file_name)
            inserted += 1
            product_key = data_service._generate_product_key(product_name, product.specification)
            # 同步一份到 Supabase Postgres（最小侵入：best effort，不影响本地 DB）
            try:
                note_parts = [file_name]
                if row_sheet:
                    note_parts.append(f"sheet={row_sheet}")
                insert_oos_record(
                    issue_type="oos",
                    product_name=product_name,
                    product_code=product_key,
                    customer_name=None,
                    need_qty=qty,
                    avail_qty=0.0,
                    shortfall_qty=qty,
                    note=" | ".join(note_parts),
                )
            except Exception as e:
                logger.warning("同步无货记录到 Supabase 失败（忽略，不影响本地写入）: %s", e)
            if data_service.should_trigger_email(product_key, count):
                from .email_service import send_out_of_stock_alert
                if send_out_of_stock_alert(
                    product_name=product_name,
                    specification=product.specification,
                    product_key=product_key,
                    count=count,
                    file_name=file_name,
                ):
                    data_service.mark_email_sent(product_key)
        except Exception as e:
            errors.append(product_name[:20] + ": " + str(e))
    message = f"已持久化 {inserted} 条。"
    if errors:
        message += " 失败: " + "; ".join(errors[:5])
    return {"success": True, "result": {"inserted": inserted, "message": message}}


def get_quotation_tools_openai_format() -> list[dict]:
    """Agent 可用工具（OpenAI function calling 格式）。"""
    return [
        {
            "type": "function",
            "function": {
                "name": "analyze_quotation_excel",
                "description": "【第一步】扫描报价单 Excel，统计「无货/缺货」关键词的行数及行号，快速了解无货规模。在 get_out_of_stock_records 之前调用。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "报价单 Excel 的完整路径或相对路径"},
                        "sheet_name": {"type": "string", "description": "指定工作表名称，不传则扫描所有 Sheet"},
                    },
                    "required": ["file_path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "get_out_of_stock_records",
                "description": "【第二步】抓取所有无货行的原始单元格数据及表头，供你观察列含义（product_name/specification/unit/quantity 各对应哪列），然后调用 persist_out_of_stock_records 落库。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {"type": "string", "description": "报价单 Excel 的完整路径或相对路径"},
                        "sheet_name": {"type": "string", "description": "指定工作表名称，不传则扫描所有 Sheet"},
                    },
                    "required": ["file_path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "persist_out_of_stock_records",
                "description": "【第三步】将无货产品记录落库（显示在无货产品列表）。records 每条填 product_name（必填）、specification、unit、quantity，跳过表头/合计/空行等非产品行。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_name": {"type": "string", "description": "原始 Excel 文件名，入库后显示在无货产品列表的 file_name 列"},
                        "sheet_name": {"type": "string", "description": "工作表名称（如「询价单」），入库后显示在 sheet_name 列；不传则为空"},
                        "records": {
                            "type": "array",
                            "description": "要存库的记录列表，每条与无货产品列表行格式一致",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "product_name": {"type": "string", "description": "产品名称（必填）"},
                                    "specification": {"type": "string", "description": "规格，如 50cm、DN100"},
                                    "unit": {"type": "string", "description": "单位，如 根、个"},
                                    "quantity": {"type": "number", "description": "数量"},
                                    "sheet_name": {"type": "string", "description": "可选，单条记录的工作表名，不传则用顶层 sheet_name"},
                                },
                                "required": ["product_name"],
                            },
                        },
                    },
                    "required": ["file_name", "records"],
                },
            },
        },
    ]


def _execute_quotation_tool_impl(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    """实际执行逻辑，供带超时调用。返回 { success, result?, error? }。"""
    if name == "analyze_quotation_excel":
        out = analyze_quotation_excel(
            arguments.get("file_path", ""),
            arguments.get("sheet_name"),
        )
        if out.get("success"):
            return {"success": True, "result": out.get("result", out)}
        return {"success": False, "error": out.get("error", "unknown")}
    if name == "get_out_of_stock_records":
        out = get_out_of_stock_records(
            arguments.get("file_path", ""),
            arguments.get("sheet_name"),
        )
        if out.get("success"):
            return {"success": True, "result": out.get("result", out)}
        return {"success": False, "error": out.get("error", "unknown")}
    if name == "persist_out_of_stock_records":
        recs = arguments.get("records")
        if not isinstance(recs, list):
            recs = []
        out = persist_out_of_stock_records(
            arguments.get("file_name", ""),
            recs,
            sheet_name=arguments.get("sheet_name", ""),
        )
        if out.get("success"):
            return {"success": True, "result": out.get("result", out)}
        return {"success": False, "error": out.get("error", "unknown")}
    return {"success": False, "error": f"Unknown tool: {name}"}


def execute_quotation_tool(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
    """执行报价工具，带超时封装（与 inventory_agent_tools 对齐）。返回 { success, result?, error? }。"""
    try:
        from backend.tools.oos.config import TOOL_EXEC_TIMEOUT
        timeout_sec = TOOL_EXEC_TIMEOUT
    except ImportError:
        timeout_sec = 35
    try:
        with ThreadPoolExecutor(max_workers=1) as ex:
            future = ex.submit(_execute_quotation_tool_impl, name, arguments)
            return future.result(timeout=timeout_sec)
    except FuturesTimeoutError:
        return {
            "success": False,
            "error": "timeout",
            "result": f"工具执行超时（{timeout_sec} 秒）。请检查文件路径与网络，或稍后重试。",
        }
    except Exception as e:
        logger.exception("quotation tool 执行异常")
        return {"success": False, "error": str(e), "result": str(e)}
