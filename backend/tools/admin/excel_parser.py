"""
上传 Excel → dict 列表（与现有万鼎/映射表列布局一致）。
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any

logger = logging.getLogger(__name__)

_COL_MATERIAL = 1
_COL_DESC = 2
_COL_PRICE_A = 8
_COL_PRICE_B = 10
_COL_PRICE_C = 12
_COL_PRICE_D = 14


def _safe_float(v: Any) -> Any:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _val(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_price_library(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl 未安装") from e

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[dict] = []

    def _load_sheet(ws):
        for i, row in enumerate(ws.iter_rows(max_col=16)):
            if i == 0:
                continue
            cells = [None] * 16
            for col_idx, c in enumerate(row):
                if col_idx < 16:
                    cells[col_idx] = getattr(c, "value", None)
            material = _val(cells[_COL_MATERIAL]) if len(cells) > _COL_MATERIAL else ""
            desc = _val(cells[_COL_DESC]) if len(cells) > _COL_DESC else ""
            if not material and not desc:
                continue
            rows.append(
                {
                    "material": material,
                    "description": desc,
                    "price_a": _safe_float(cells[_COL_PRICE_A]) if len(cells) > _COL_PRICE_A else None,
                    "price_b": _safe_float(cells[_COL_PRICE_B]) if len(cells) > _COL_PRICE_B else None,
                    "price_c": _safe_float(cells[_COL_PRICE_C]) if len(cells) > _COL_PRICE_C else None,
                    "price_d": _safe_float(cells[_COL_PRICE_D]) if len(cells) > _COL_PRICE_D else None,
                }
            )

    ws_main = wb["管材"] if "管材" in wb.sheetnames else (wb.active or wb[wb.sheetnames[0]])
    _load_sheet(ws_main)
    if "国标管件" in wb.sheetnames:
        _load_sheet(wb["国标管件"])
    wb.close()
    logger.info("parse_price_library: %d rows", len(rows))
    return rows


def parse_product_mapping(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl 未安装") from e

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(max_col=4, min_row=2)):
        cells = [getattr(c, "value", None) for c in row]
        if len(cells) < 3:
            continue
        product_code = _val(cells[2]) if len(cells) > 2 else ""
        if not product_code:
            continue
        rows.append(
            {
                "inquiry_name": _val(cells[0]),
                "spec": _val(cells[1]) if len(cells) > 1 else "",
                "product_code": product_code,
                "quotation_name": _val(cells[3]) if len(cells) > 3 else "",
            }
        )
    wb.close()
    logger.info("parse_product_mapping: %d rows", len(rows))
    return rows


def _sanitize_col_name(name: str) -> str:
    """Normalize column names to safe identifiers for dynamic SQL tables."""
    clean = name.replace("\x00", "").replace('"', "").strip()
    clean = re.sub(r"[\s\-().;,/\\]+", "_", clean)
    clean = re.sub(r"_+", "_", clean).strip("_")
    return clean or "col"


def _make_unique_names(raw_names: list[str]) -> tuple[list[str], list[str]]:
    warnings: list[str] = []
    seen: dict[str, int] = {}
    used: set[str] = set()
    result: list[str] = []
    for i, name in enumerate(raw_names, start=1):
        n = (name or "").strip()
        if not n:
            n = f"col_{i}"
            warnings.append(f'第 {i} 列空列头，已自动命名为 "{n}"')
        seen[n] = seen.get(n, 0) + 1
        if seen[n] > 1:
            candidate_index = seen[n]
            new_n = f"{n}_{candidate_index}"
            while new_n in used:
                candidate_index += 1
                new_n = f"{n}_{candidate_index}"
            warnings.append(f'列头 "{n}" 重复，第 {i} 列重命名为 "{new_n}"')
            n = new_n
            seen[n] = seen.get(n, 0) + 1
        if n in used:
            candidate_index = 2
            new_n = f"{n}_{candidate_index}"
            while new_n in used:
                candidate_index += 1
                new_n = f"{n}_{candidate_index}"
            warnings.append(f'列头 "{n}" 重复，第 {i} 列重命名为 "{new_n}"')
            n = new_n
        result.append(n)
        used.add(n)
    return result, warnings


def _infer_type(values: list[Any]) -> tuple[str, list[str]]:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "TEXT", []
    numeric_count = 0
    for v in non_empty:
        try:
            float(str(v).replace(",", ""))
            numeric_count += 1
        except (ValueError, TypeError):
            continue
    ratio = numeric_count / len(non_empty)
    if ratio >= 0.9:
        return "NUMERIC", []
    if 0 < ratio < 0.9:
        return "TEXT", [f"类型混乱（{numeric_count}/{len(non_empty)} 可解析为数字），已降级为 TEXT"]
    return "TEXT", []


def _parse_rows_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl 未安装") from e

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return [], []
    headers = [str(v) if v is not None else "" for v in all_rows[0]]
    data = [list(row) for row in all_rows[1:]]
    return headers, data


def _parse_rows_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    import csv as csv_mod

    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv_mod.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv_mod.Error:
        dialect = csv_mod.excel

    reader = csv_mod.reader(io.StringIO(text), dialect)
    all_rows = list(reader)
    if not all_rows:
        return [], []
    headers = all_rows[0]
    data = [row for row in all_rows[1:] if any(v.strip() for v in row)]
    return headers, data


def parse_generic(content: bytes, filename: str) -> dict:
    """
    Parse arbitrary Excel/CSV and return detected schema + aligned rows.
    """
    errors: list[str] = []
    global_warnings: list[str] = []

    try:
        ext = (filename or "").rsplit(".", 1)[-1].lower()
        if ext == "xlsx":
            raw_headers, raw_rows = _parse_rows_xlsx(content)
        elif ext == "xls":
            return {
                "columns": [],
                "rows": [],
                "warnings": [],
                "errors": ["不支持 .xls 格式，请另存为 .xlsx 后重新上传"],
            }
        else:
            raw_headers, raw_rows = _parse_rows_csv(content)
    except Exception as e:
        return {"columns": [], "rows": [], "warnings": [], "errors": [f"文件解析失败: {e}"]}

    if len(raw_headers) > 100:
        errors.append(f"列数超过 100（当前 {len(raw_headers)} 列），请拆分文件后重新上传")
        return {"columns": [], "rows": [], "warnings": [], "errors": errors}

    data_rows = [r for r in raw_rows if any(v is not None and str(v).strip() != "" for v in r)]
    if not data_rows:
        errors.append("文件没有数据行（仅有表头或文件为空）")
        return {"columns": [], "rows": [], "warnings": [], "errors": errors}

    # 过滤掉原始列头为空/None 的列，避免生成 col_N 占位列
    non_empty_col_indices = [i for i, h in enumerate(raw_headers) if h and h.strip()]
    if len(non_empty_col_indices) < len(raw_headers):
        dropped_count = len(raw_headers) - len(non_empty_col_indices)
        raw_headers = [raw_headers[i] for i in non_empty_col_indices]
        data_rows = [
            [row[i] if i < len(row) else None for i in non_empty_col_indices]
            for row in data_rows
        ]
        global_warnings.append(f"已过滤 {dropped_count} 个空列头列")
    if not raw_headers:
        errors.append("所有列头均为空，无法导入")
        return {"columns": [], "rows": [], "warnings": global_warnings, "errors": errors}

    ncols = len(raw_headers)
    aligned: list[list[Any]] = []
    for row in data_rows:
        if len(row) < ncols:
            aligned.append(list(row) + [None] * (ncols - len(row)))
        else:
            aligned.append(list(row[:ncols]))

    unique_names, name_warnings = _make_unique_names(raw_headers)
    global_warnings.extend(name_warnings)

    col_defs: list[dict[str, Any]] = []
    used_sanitized: set[str] = set()
    scan_limit = min(200, len(aligned))
    for ci, col_name in enumerate(unique_names):
        col_values = [aligned[ri][ci] for ri in range(scan_limit)]
        col_type, type_warnings = _infer_type(col_values)
        col_warnings: list[str] = list(type_warnings)

        non_empty = sum(1 for v in col_values if v is not None and str(v).strip() != "")
        if scan_limit > 0 and non_empty / scan_limit < 0.2:
            warning = f'列 "{col_name}" 超过 80% 的空值'
            col_warnings.append(warning)
            global_warnings.append(warning)

        sanitized = _sanitize_col_name(col_name)
        if sanitized in used_sanitized:
            suffix = 2
            candidate = f"{sanitized}_{suffix}"
            while candidate in used_sanitized:
                suffix += 1
                candidate = f"{sanitized}_{suffix}"
            global_warnings.append(f'列 "{col_name}" 清洗后重名，已重命名为 "{candidate}"')
            sanitized = candidate
        used_sanitized.add(sanitized)

        col_defs.append(
            {
                "name": sanitized,
                "type": col_type,
                "original_name": raw_headers[ci],
                "warnings": col_warnings,
            }
        )

    for col in col_defs:
        for warning in col["warnings"]:
            if warning not in global_warnings:
                global_warnings.append(warning)

    return {
        "columns": col_defs,
        "rows": aligned,
        "warnings": global_warnings,
        "errors": [],
    }
