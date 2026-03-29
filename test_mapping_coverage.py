"""
测试报价历史产品在字段匹配中的覆盖率
读取 整理产品(2).xlsx，对每一行调用字段匹配，判断是否能包含该行
"""
import asyncio
import sys
import os
import json
import re

# 确保项目根目录在 path 中
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed, trying pandas")
    import pandas as pd

async def load_mapping_table(path):
    """加载映射表 Excel"""
    rows = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(max_col=4, min_row=2)):
            cells = [getattr(c, 'value', None) for c in row]
            if len(cells) < 4:
                continue
            field_name = str(cells[0] or "").strip()
            spec = str(cells[1] or "").strip() if cells[1] is not None else ""
            code = cells[2]
            if code is None:
                continue
            code = str(code).strip()
            matched_name = str(cells[3] or "").strip()
            if not field_name and not code:
                continue
            search_text = f"{field_name} {spec}".strip() if spec else field_name
            rows.append({
                "search_text": search_text,
                "code": code,
                "matched_name": matched_name,
                "spec": spec,
            })
        wb.close()
        return rows
    except Exception as e:
        print(f"Error loading mapping table: {e}")
        return []


def normalize(s):
    """规范化文本"""
    s = (s or "").lower().strip()
    s = s.replace("－", "-").replace("—", "-").replace("（", "(").replace("）", ")")
    s = re.sub(r"[_\t]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def split_tokens(text):
    """简单 token 分割"""
    text = normalize(text)
    tokens = []
    for m in re.finditer(r"[\u4e00-\u9fff]+|\d+(?:\.\d+)?", text):
        tokens.append(m.group())
    return tokens


async def match_wanding_price_candidates(keywords, customer_level="B"):
    """调用字段匹配"""
    try:
        from backend.tools.inventory.services.match_and_inventory import match_wanding_price_candidates as do_match
        candidates = do_match(keywords, customer_level=customer_level, max_score_tiers=2)
        return candidates
    except Exception as e:
        print(f"Match error for '{keywords}': {e}")
        return []


def check_coverage(row, candidates):
    """检查候选中是否包含该行的 code 或 matched_name"""
    code = row["code"]
    matched_name = row["matched_name"]

    for c in candidates:
        c_code = str(c.get("code", "")).strip()
        c_name = str(c.get("matched_name", "")).strip()

        if c_code == code:
            return "code_match"
        # 模糊匹配：报价名称包含在候选名称中，或候选名称包含在报价名称中
        if matched_name and c_name:
            if matched_name in c_name or c_name in matched_name:
                return "name_match"
            # 归一化后匹配
            if normalize(matched_name) in normalize(c_name) or normalize(c_name) in normalize(matched_name):
                return "name_match"

    return None


async def process_all():
    """处理所有行"""
    path = r"d:\Projects\agent-jk\Agent Team version3\data\整理产品(2).xlsx"
    print(f"Loading mapping table from: {path}")
    rows = await load_mapping_table(path)
    print(f"Total rows loaded: {len(rows)}")

    if not rows:
        print("No rows loaded, exiting")
        return

    covered = []
    not_covered = []

    for i, row in enumerate(rows):
        keywords = row["search_text"]
        code = row["code"]
        matched_name = row["matched_name"]

        if not keywords:
            continue

        # 调用字段匹配
        candidates = await match_wanding_price_candidates(keywords)

        # 检查覆盖率
        result = check_coverage(row, candidates)

        record = {
            "index": i + 2,  # Excel 行号（含表头）
            "search_text": keywords,
            "code": code,
            "matched_name": matched_name,
            "candidates_count": len(candidates),
            "candidates": [{"code": c.get("code"), "matched_name": c.get("matched_name")[:50]} for c in candidates[:5]],
        }

        if result:
            record["match_type"] = result
            covered.append(record)
        else:
            record["match_type"] = "not_found"
            not_covered.append(record)

        # 进度显示
        if (i + 1) % 20 == 0:
            print(f"Processed {i + 1}/{len(rows)} rows, covered: {len(covered)}, not covered: {len(not_covered)}")

    # 输出结果
    print("\n" + "=" * 80)
    print(f"COVERAGE REPORT")
    print(f"Total rows: {len(rows)}")
    print(f"Covered: {len(covered)} ({100*len(covered)/len(rows):.1f}%)")
    print(f"Not covered: {len(not_covered)} ({100*len(not_covered)/len(rows):.1f}%)")
    print("=" * 80)

    # 保存详细结果
    result_data = {
        "summary": {
            "total": len(rows),
            "covered": len(covered),
            "not_covered": len(not_covered),
            "coverage_rate": f"{100*len(covered)/len(rows):.1f}%",
        },
        "covered": covered,
        "not_covered": not_covered,
    }

    output_path = r"d:\Projects\agent-jk\Agent Team version3\data\coverage_report.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result_data, f, ensure_ascii=False, indent=2)
    print(f"\nDetailed report saved to: {output_path}")

    # 打印不能包含的列表
    if not_covered:
        print("\n" + "=" * 80)
        print("NOT COVERED ITEMS:")
        print("=" * 80)
        for item in not_covered[:30]:  # 先显示前30个
            print(f"  Row {item['index']}: '{item['search_text']}'")
            print(f"    Code: {item['code']}, Name: {item['matched_name']}")
            if item['candidates']:
                print(f"    Top candidates:")
                for c in item['candidates'][:3]:
                    print(f"      - {c['code']}: {c['matched_name']}")
            else:
                print(f"    No candidates returned!")
            print()
        if len(not_covered) > 30:
            print(f"  ... and {len(not_covered) - 30} more not covered items")


if __name__ == "__main__":
    asyncio.run(process_all())
