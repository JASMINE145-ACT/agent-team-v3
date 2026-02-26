# -*- coding: utf-8 -*-
"""
通过相同 code 用 item-list-slim 的英文名补全价格库的 Describrition_English（D 列）。
- 价格库：万鼎价格库_管材与国标管件_标准格式.xlsx，B 列 = Material（编码），D 列 = Describrition_English
- 清单：item-list-slim.xlsx，A 列 = Item Code，B 列 = Item Name（英文）
- 匹配键：Material 与 Item Code 转为字符串后一致即视为相同 code
- 输出：同目录下 万鼎价格库_管材与国标管件_标准格式_with_english.xlsx（保留原表公式，仅填充/覆盖 D 列）
"""
from __future__ import annotations

import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("需要 openpyxl: pip install openpyxl")

# 路径（脚本放在 data 目录下，与价格库同目录）
DATA_DIR = Path(__file__).resolve().parent
PRICE_LIB_PATH = DATA_DIR / "万鼎价格库_管材与国标管件_标准格式.xlsx"
ITEM_LIST_PATH = Path(__file__).resolve().parents[2] / "数据标准化" / "02_待处理数据" / "item-list-slim.xlsx"
OUT_PATH = DATA_DIR / "万鼎价格库_管材与国标管件_标准格式_with_english.xlsx"


def _norm_code(v) -> str:
    """编码统一为字符串便于匹配。"""
    if v is None:
        return ""
    s = str(v).strip()
    # 若 Excel 读成浮点如 8010012683.0，去掉 .0
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def build_code_to_english(item_list_path: Path) -> dict[str, str]:
    """从 item-list-slim 读取 Item Code -> Item Name 映射。"""
    wb = openpyxl.load_workbook(item_list_path, read_only=True, data_only=True)
    ws = wb.active
    code_to_english = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        code_val, name_val = row[0], row[1]
        code = _norm_code(code_val)
        if not code:
            continue
        name = name_val if name_val is not None else ""
        if isinstance(name, (int, float)) and not isinstance(name, bool):
            name = str(int(name)) if isinstance(name, float) and name == int(name) else str(name)
        else:
            name = str(name).strip() if name else ""
        # 若同一 code 出现多次，后者覆盖（一般 item-list 每 code 一行）
        code_to_english[code] = name
    wb.close()
    return code_to_english


def fill_price_lib_english(
    price_lib_path: Path,
    out_path: Path,
    code_to_english: dict[str, str],
) -> tuple[int, int]:
    """
    打开价格库，按 B 列 Material 匹配 code，将对应英文写入 D 列；保存到 out_path。
    返回 (匹配并写入行数, 总数据行数)。
    """
    wb = openpyxl.load_workbook(price_lib_path, data_only=False)  # 保留公式
    written = 0
    total_data_rows = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # D 列 = 4
        for row in range(2, ws.max_row + 1):
            code_val = ws.cell(row=row, column=2).value  # B
            if code_val is None and ws.cell(row=row, column=3).value is None:
                # 空行可跳过
                continue
            total_data_rows += 1
            code = _norm_code(code_val)
            if not code:
                continue
            en = code_to_english.get(code)
            if en is None:
                continue
            ws.cell(row=row, column=4).value = en  # D 列
            written += 1
    wb.save(out_path)
    return written, total_data_rows


def main() -> None:
    if not ITEM_LIST_PATH.exists():
        raise SystemExit(f"item-list 不存在: {ITEM_LIST_PATH}")
    if not PRICE_LIB_PATH.exists():
        raise SystemExit(f"价格库不存在: {PRICE_LIB_PATH}")

    code_to_english = build_code_to_english(ITEM_LIST_PATH)
    print(f"item-list 中编码数量: {len(code_to_english)}")

    written, total = fill_price_lib_english(PRICE_LIB_PATH, OUT_PATH, code_to_english)
    print(f"价格库数据行数: {total}，按 code 匹配并写入英文: {written} 行")
    print(f"已保存: {OUT_PATH}")


if __name__ == "__main__":
    main()
