"""
报价链路近期修改的简单测试：Match 接收 items、缺货不阻塞、fill 日期/规格回退、上传 with_summary 快路径。
运行（在 Agent Team version3 根目录）：
  python -m pytest tests/test_quotation_pipeline.py -v
"""
import json
import sys
import tempfile
import time
from pathlib import Path

_root = Path(__file__).resolve().parent.parent
if str(_root) not in sys.path:
    sys.path.insert(0, str(_root))

import pytest


# ----- 1) Match 接收已提取的 items（去重 extract） -----


def test_match_uses_items_when_provided_no_extract_call():
    """传入合法 items 时，直接使用 items、不再次调用 extract_inquiry_items。"""
    from unittest.mock import patch, MagicMock
    from backend.agent.work_tools import _run_work_quotation_match

    items = [
        {"row": 2, "keywords": "PVC管 dn20", "qty": 100, "product_name": "PVC管", "specification": "dn20"},
        {"row": 3, "keywords": "阀门", "qty": 50, "product_name": "阀门", "specification": ""},
    ]
    # 模拟 match_price_and_get_inventory：第一条有库存，第二条缺货
    def fake_match(keywords, customer_level=None, price_library_path=None, allow_suggestions_for_work=None):
        if "PVC" in (keywords or ""):
            return {"code": "P001", "matched_name": "PVC管", "unit_price": 10.0, "available_qty": 200.0}
        return {"code": "V001", "matched_name": "阀门", "unit_price": 20.0, "available_qty": 10.0}

    with patch("backend.tools.quotation.quote_tools.extract_inquiry_items") as mock_extract:
        with patch("backend.tools.inventory.services.match_and_inventory.match_price_and_get_inventory", side_effect=fake_match):
            out = _run_work_quotation_match("/any/file.xlsx", items=items)

    mock_extract.assert_not_called()
    assert out.get("success") is True
    assert "to_fill" in out and "shortage" in out and "unmatched" in out
    assert len(out.get("items", [])) == 2
    assert len(out.get("to_fill", [])) == 1
    assert len(out.get("shortage", [])) == 1
    assert out["to_fill"][0].get("row") == 2 and out["shortage"][0].get("row") == 3


def test_match_items_structure_to_fill_shortage_unmatched():
    """传入 items 时返回的 to_fill/shortage/unmatched/fill_items_merged 结构正确。"""
    from unittest.mock import patch
    from backend.agent.work_tools import _run_work_quotation_match

    items = [{"row": 2, "keywords": "A", "qty": 1, "product_name": "A", "specification": ""}]
    with patch("backend.tools.inventory.services.match_and_inventory.match_price_and_get_inventory") as m:
        m.return_value = {"code": "C1", "matched_name": "A", "unit_price": 1.0, "available_qty": 10.0}
        out = _run_work_quotation_match("/f.xlsx", items=items)

    assert out["success"] is True
    assert len(out["to_fill"]) == 1
    assert out["to_fill"][0]["code"] == "C1"
    assert len(out["shortage"]) == 0
    assert len(out["unmatched"]) == 0
    assert "fill_items_merged" in out
    assert len(out["fill_items_merged"]) >= 1


# ----- 2) 缺货落库/邮件不阻塞主流程 -----


def test_execute_work_tool_sync_returns_quickly_with_shortage():
    """execute_work_tool_sync 在返回含 shortage 的 result 时立即返回，不等待落库。"""
    from unittest.mock import patch
    from backend.agent.work_tools import execute_work_tool_sync

    block_called = []

    def blocking_persist(*args, **kwargs):
        block_called.append(1)
        time.sleep(2.0)

    # 模拟 match 返回缺货，避免真实请求耗时；落库在后台线程中若未 mock 会阻塞 2s，主线程不应等它
    def fake_match_shortage(*args, **kwargs):
        return {"code": "C1", "matched_name": "X", "unit_price": 1.0, "available_qty": 5.0}

    with patch("backend.agent.work_tools._persist_shortage_records_and_alerts", side_effect=blocking_persist):
        with patch("backend.tools.inventory.services.match_and_inventory.match_price_and_get_inventory", side_effect=fake_match_shortage):
            with patch("backend.tools.quotation.spec_extract.extract_specs_batch_llm", return_value=[]):
                t0 = time.perf_counter()
                raw = execute_work_tool_sync(
                "work_quotation_match",
                {
                    "file_path": "/f.xlsx",
                    "items": [{"row": 2, "keywords": "X", "qty": 100, "product_name": "X", "specification": ""}],
                },
                )
                elapsed = time.perf_counter() - t0

    # 主线程应在 1.5 秒内返回（match 已 mock，落库在后台 daemon 线程不阻塞主线程；规范行构建内 batch LLM 已 mock）
    assert elapsed < 1.5, "主流程不应等待落库，应在 1.5 秒内返回"
    data = json.loads(raw)
    assert data.get("success") is True
    assert "shortage" in data
    # 后台线程可能已启动，但不影响主线程已返回
    assert "to_fill" in data and "fill_items_merged" in data


# ----- 3) fill_quotation 填写日期与规格回退 -----


@pytest.fixture
def temp_quotation_xlsx():
    """生成带「交货日期」「报价日期」表头和 Total 行的临时 xlsx。"""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sheet1")
    ws.sheet_view.view = "pageBreakPreview"
    # 表头：第1行 含 交货日期；合计行下方含 报价日期
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=7, value="产品编号")
    ws.cell(row=1, column=8, value="报价名称")
    ws.cell(row=1, column=10, value="规格")
    ws.cell(row=1, column=12, value="数量")
    ws.cell(row=1, column=14, value="单价")
    ws.cell(row=1, column=15, value="总价")
    delivery_col = 16
    ws.cell(row=1, column=delivery_col, value="交货日期")
    # 数据行
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=3, column=1, value=2)
    # Total Excluding PPN 行（用于 quote_tools 定位）
    total_row = 5
    ws.cell(row=total_row, column=1, value="Total Excluding PPN不含税总价")
    ws.cell(row=total_row + 4, column=1, value="报价日期")
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    import os
    os.close(fd)
    wb.save(path)
    yield path
    try:
        Path(path).unlink(missing_ok=True)
    except Exception:
        pass


def test_fill_quotation_writes_dates_and_spec_fallback(temp_quotation_xlsx):
    """fill_quotation 写入报价日期/交货日期；无 specification 时规格列用 quote_name。"""
    from backend.tools.quotation.quote_tools import fill_quotation
    import openpyxl

    fill_items = [
        {"row": 2, "code": "C1", "quote_name": "产品A", "unit_price": 10.0, "qty": 2, "specification": ""},
        {"row": 3, "code": "C2", "quote_name": "产品B", "unit_price": 5.0, "qty": 4, "specification": "型号X"},
    ]
    out_path = temp_quotation_xlsx + ".filled.xlsx"
    out = fill_quotation(
        file_path=temp_quotation_xlsx,
        fill_items=fill_items,
        output_path=out_path,
        quotation_date="2026/03/10",
        delivery_date="2026/03/20",
    )
    assert out.get("success") is True
    assert Path(out_path).exists()

    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.active
    assert ws.sheet_view.view == "normal"
    # 规格列 J=10：第2行无 specification 应写 quote_name「产品A」；第3行有 specification 写「型号X」
    spec_2 = ws.cell(row=2, column=10).value
    spec_3 = ws.cell(row=3, column=10).value
    assert (spec_2 or "").strip() == "产品A"
    assert (spec_3 or "").strip() == "型号X"
    # 交货日期列：表头在 16，数据行应写入 2026/03/20（需确认 _find_delivery_date_column 找到的是第几列）
    # 本 fixture 把「交货日期」写在 column 16，所以 delivery_col=16，第2、3行该列应有 2026/03/20
    delivery_val = ws.cell(row=2, column=16).value
    assert delivery_val == "2026/03/20" or "2026" in str(delivery_val)

    # 至少验证 1 个非数据列（不在 G/H/J/L/N/O 中）的样式与模板行保持一致，避免样式复制仅覆盖少数列。
    # 这里选用第 5 列（E 列），它在 STYLE_COL_START..STYLE_COL_END_DEFAULT 范围内，但业务上并非关键数据列。
    template_row = 1  # fill_quotation 选用首个数据行之上一行为模板行，本 fixture 的数据从第 2 行开始
    data_row = 2
    non_data_col = 5
    tmpl_cell = ws.cell(row=template_row, column=non_data_col)
    filled_cell = ws.cell(row=data_row, column=non_data_col)
    # 不依赖具体颜色/边框样式，只要求复制后 border/fill 的序列化结果一致（避免因对象实例不同导致比较失败）
    assert repr(tmpl_cell.border) == repr(filled_cell.border)
    assert repr(tmpl_cell.fill) == repr(filled_cell.fill)
    wb.close()
    try:
        Path(out_path).unlink(missing_ok=True)
    except Exception:
        pass


# ----- 4) 上传 with_summary 快路径 -----


def test_upload_with_summary_zero_returns_no_summary_meta():
    """POST /api/quotation/upload?with_summary=0 时响应不含 summary_meta 且 200。"""
    from fastapi.testclient import TestClient
    from backend.server.api.app import app
    import io

    client = TestClient(app)
    # 最小 xlsx：openpyxl 新建空工作簿
    try:
        import openpyxl
        buf = io.BytesIO()
        openpyxl.Workbook().save(buf)
        buf.seek(0)
        content = buf.getvalue()
    except ImportError:
        pytest.skip("openpyxl not installed")

    resp = client.post(
        "/api/quotation/upload",
        params={"with_summary": 0},
        files={"file": ("small.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data.get("success") is True
    assert "data" in data
    # with_summary=0 时不应包含 summary_meta（或 summary_meta 为空）
    inner = data.get("data") or {}
    assert "summary_meta" not in inner or inner.get("summary_meta") is None


def test_fill_template_with_inquiry_items_normalizes_sheet_view(tmp_path):
    """Generated inquiry file should be opened in normal view, not page-break preview."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")
    from backend.tools.quotation.quote_tools import fill_template_with_inquiry_items

    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询价单"
    ws.sheet_view.view = "pageBreakPreview"
    ws.cell(row=1, column=2, value="询价货物名称")
    ws.cell(row=1, column=3, value="询价规格型号")
    ws.cell(row=1, column=5, value="数量")
    ws.cell(row=9, column=1, value="Total Excluding PPN不含税总价")
    wb.save(template)
    wb.close()

    out = fill_template_with_inquiry_items(
        template_path=str(template),
        items=[{"product_name": "A", "specification": "x", "qty": 1}],
        output_path=str(output),
        sheet_name="询价单",
        allow_insert_rows=False,
    )
    assert out.get("success") is True
    assert output.exists()

    out_wb = openpyxl.load_workbook(output)
    out_ws = out_wb["询价单"]
    assert out_ws.sheet_view.view == "normal"
    out_wb.close()


# ----- 5) 规范行单源：build_canonical_quotation_lines / fill_items_from_canonical_lines -----


def test_build_canonical_quotation_lines_structure():
    """build_canonical_quotation_lines 产出规范行，含 specification/quote_spec 等，且不调 LLM 时规格来自规则。"""
    from backend.tools.quotation.canonical_lines import build_canonical_quotation_lines

    fill_items_merged = [
        {"row": 2, "code": "C1", "quote_name": "PVC管 DN20", "unit_price": 10.0, "qty": 100, "specification": "dn20"},
        {"row": 3, "code": "无货", "quote_name": "", "unit_price": None, "qty": 50, "specification": ""},
    ]
    items = [
        {"row": 2, "product_name": "PVC管", "specification": "dn20", "qty": 100},
        {"row": 3, "product_name": "阀门", "specification": "", "qty": 50},
    ]
    shortage = [{"row": 2, "shortfall": 20.0, "available_qty": 80.0}]
    lines = build_canonical_quotation_lines(
        fill_items_merged, items, shortage, run_spec_llm=False
    )
    assert len(lines) == 2
    for line in lines:
        for key in ("row", "row_index", "product_name", "specification", "qty", "code", "quote_name", "quote_spec",
                    "unit_price", "amount", "available_qty", "shortfall", "is_shortage", "match_source"):
            assert key in line
    assert lines[0]["row"] == 2 and lines[0]["code"] == "C1"
    assert lines[0]["specification"] == "dn20"
    assert lines[0]["quote_spec"]  # 从 quote_name "PVC管 DN20" 规则抽
    assert lines[0]["shortfall"] == 20.0 and lines[0]["available_qty"] == 80.0
    assert lines[1]["code"] == "无货" and lines[1]["is_shortage"] == 1
    assert lines[1]["shortfall"] == 50 and lines[1]["available_qty"] == 0.0


def test_fill_items_from_canonical_lines():
    """fill_items_from_canonical_lines 导出 fill_quotation 入参，specification 取 quote_spec 或 specification。"""
    from backend.tools.quotation.canonical_lines import fill_items_from_canonical_lines

    canonical_lines = [
        {"row": 2, "code": "C1", "quote_name": "A", "unit_price": 1.0, "qty": 10,
         "specification": "询价规", "quote_spec": "报价规"},
        {"row": 3, "code": "C2", "quote_name": "B", "unit_price": 2.0, "qty": 5,
         "specification": "仅询价", "quote_spec": ""},
    ]
    fill_items = fill_items_from_canonical_lines(canonical_lines)
    assert len(fill_items) == 2
    assert fill_items[0]["specification"] == "报价规"
    assert fill_items[1]["specification"] == "仅询价"
    assert fill_items[0]["row"] == 2 and fill_items[0]["code"] == "C1" and fill_items[0]["qty"] == 10


def test_match_output_includes_fill_items_for_excel():
    """work_quotation_match 在 success 且非 needs_human_choice 时产出 fill_items_for_excel 与 pending_quotation_draft。"""
    from unittest.mock import patch
    from backend.agent.work_tools import execute_work_tool_sync

    with patch("backend.tools.inventory.services.match_and_inventory.match_price_and_get_inventory") as m:
        m.return_value = {"code": "C1", "matched_name": "X", "unit_price": 1.0, "available_qty": 100.0}
        raw = execute_work_tool_sync(
            "work_quotation_match",
            {"file_path": "/f.xlsx", "items": [{"row": 2, "keywords": "X", "qty": 10, "product_name": "X", "specification": ""}]},
        )
    data = json.loads(raw)
    assert data.get("success") is True
    assert data.get("needs_human_choice") is not True
    assert "fill_items_for_excel" in data
    assert "pending_quotation_draft" in data
    assert "lines" in data["pending_quotation_draft"]
    assert len(data["fill_items_for_excel"]) == len(data["pending_quotation_draft"]["lines"])


def test_build_canonical_quotation_lines_handles_non_numeric_values():
    """Non-numeric qty/unit_price/shortage fields should not crash canonical line build."""
    from backend.tools.quotation.canonical_lines import build_canonical_quotation_lines

    fill_items_merged = [
        {"row": 2, "code": "C1", "quote_name": "A", "unit_price": "N/A", "qty": "abc", "specification": ""},
    ]
    items = [{"row": 2, "product_name": "A", "specification": "", "qty": "x"}]
    shortage = [{"row": 2, "shortfall": "bad", "available_qty": "unknown"}]

    lines = build_canonical_quotation_lines(fill_items_merged, items, shortage, run_spec_llm=False)
    assert len(lines) == 1
    assert lines[0]["qty"] == 0.0
    assert lines[0]["unit_price"] is None
    assert lines[0]["amount"] is None
    assert lines[0]["shortfall"] == 0.0
    assert lines[0]["available_qty"] == 0.0


def test_from_text_split_mode_keeps_legacy_single_file_keys(tmp_path):
    """Split mode returns file_paths and keeps legacy file_path/file_name for compatibility."""
    from unittest.mock import patch
    from fastapi.testclient import TestClient
    from backend.server.api.app import app

    template_path = tmp_path / "template.xlsx"
    template_path.write_bytes(b"placeholder")

    fake_items = [
        {"product_name": "A", "specification": "", "qty": 1},
        {"product_name": "B", "specification": "", "qty": 2},
        {"product_name": "C", "specification": "", "qty": 3},
    ]

    with patch("backend.server.api.routes_upload.text_to_inquiry_items", return_value=fake_items):
        with patch("backend.server.api.routes_upload.get_template_inquiry_capacity", return_value={"success": True, "capacity": 2}):
            with patch("backend.server.api.routes_upload.fill_template_with_inquiry_items", return_value={"success": True}):
                client = TestClient(app)
                resp = client.post(
                    "/api/quotation/from-text",
                    json={"text": "dummy", "template_path": str(template_path)},
                )

    assert resp.status_code == 200
    body = resp.json()
    assert body.get("success") is True
    data = body.get("data") or {}
    assert data.get("chunk_count") == 2
    assert isinstance(data.get("file_paths"), list) and len(data["file_paths"]) == 2
    assert data.get("file_path") == data["file_paths"][0]
    assert data.get("file_name") == data.get("file_names", [None])[0]
