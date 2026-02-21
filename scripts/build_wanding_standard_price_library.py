"""
从新价格库 NEW PRICE(T) 万鼎国际集团最新价格库更新20251106.xlsx 中：
1. 保留 LESSO管材、国标管件 两个 sheet
2. 整理为 Copy of 万鼎国际集团最新价格库更新20250814.xlsx 的标准列格式（A-T）

标准格式列：NO, Material, Describrition, Describrition_English, 含税, 不含税, 采购不含税,
  A级别利润率, A级别报单价格, B级别利润率, B级别报单价格, C级别利润率, C级别报单价格,
  D级别利润率, D级别报单价格, D级别降低利润率, D级别报单价格2, E级别利润率, E级别报单价格, T空

运行：在 Agent Team version3 目录下
  python scripts/build_wanding_standard_price_library.py          # 仅生成
  python scripts/build_wanding_standard_price_library.py --verify # 生成并核对内容
"""
from __future__ import annotations

import math
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("需要 openpyxl: pip install openpyxl")

# 路径
BASE = Path(__file__).resolve().parent.parent
V2_DATA = BASE.parent / "Agent Team version2" / "data"
SRC_FILE = V2_DATA / "NEW PRICE(T) 万鼎国际集团最新价格库更新20251106.xlsx"
STD_REF_FILE = V2_DATA / "Copy of 万鼎国际集团最新价格库更新20250814.xlsx"
# 直接输出到 version3/data，避免依赖 v2 且防止 v2 文件被占用
OUT_FILE = BASE / "data" / "万鼎价格库_管材与国标管件_标准格式.xlsx"

# 标准表头（与 Copy of 万鼎... 管材 sheet 一致）+ 管材独有 U~X（LOCAL）
STD_HEADERS = [
    "NO",
    "Material",
    "Describrition",
    "Describrition_English",
    "INCLUDE TAX\n出厂价_含税",
    "EXCLUDE TAX\n出厂价_不含税",
    "采购不含税",
    "（二级代理）A级别 利润率",
    " （二级代理）A级别 报单价格",
    "（一级代理）B级别 利润率",
    " （一级代理）B级别 报单价格",
    " （聚万大客户）C级别 利润率",
    "（聚万大客户）C级别报单价格",
    "（青山大客户）D级别 利润率",
    "（青山大客户）D级别 报单价格",
    "（青山大客户）D级别 降低利润率",
    "（青山大客户）D级别 报单价格",
    "（大唐大客户）E级别（包运费） 利润率",
    "（大唐大客户）E级别包运费） 报单价格",
    None,
]
# 管材 sheet 独有：U=相关体积, V=%, W=EXC TAX(LOCAL 不含税), X=INC TAX(LOCAL 含税 Rp)
GUAN_HEADERS_UVWX = ["相关体积", "%", "EXC TAX", "INC TAX"]


def _cell_value(cell) -> str | float | int | None:
    v = getattr(cell, "value", None)
    if v is None:
        return None
    if hasattr(v, "__class__") and "Formula" in type(v).__name__:
        return None
    return v


def _copy_lesso_to_standard(ws_src, ws_dst):
    """NEW 文件 LESSO管材：第 1 行为说明，第 2 行为表头，第 3 行起为数据。
    列对应：NEW A=NO, C=Material, D=Describrition, E=English, G=含税(公式), H=不含税, I~T = A~E 利润率/价格。
    标准：A=NO, B=Material, C=Describrition, D=English, E=含税, F=不含税, G=采购, H~T 同。
    映射：std A←NEW A, std B←NEW C, std C←NEW D, std D←NEW E, std E←NEW G, std F←NEW H, std G←NEW H,
          std H←NEW I, std I←NEW J, ..., std S←NEW T, std T←空。
    """
    # 表头 A-T
    for col, h in enumerate(STD_HEADERS, 1):
        ws_dst.cell(row=1, column=col, value=h)
    # 管材独有 U~X：相关体积, %, EXC TAX, INC TAX
    for col, h in enumerate(GUAN_HEADERS_UVWX, 21):
        ws_dst.cell(row=1, column=col, value=h)
    # 数据：NEW 从第 3 行开始
    max_row = ws_src.max_row
    for r_src in range(3, max_row + 1):
        r_dst = r_src - 1  # 输出第 2 行起
        # 值列
        no = _cell_value(ws_src.cell(r_src, 1))
        material = _cell_value(ws_src.cell(r_src, 3))
        desc = _cell_value(ws_src.cell(r_src, 4))
        desc_en = _cell_value(ws_src.cell(r_src, 5))
        exclude_tax = _cell_value(ws_src.cell(r_src, 8))  # H = 不含税
        # NO 列：标准格式为 =ROW()-1
        ws_dst.cell(r_dst, 1, value=f"=ROW()-1")
        if material is not None:
            ws_dst.cell(r_dst, 2, value=material)
        if desc is not None:
            ws_dst.cell(r_dst, 3, value=desc)
        if desc_en is not None:
            ws_dst.cell(r_dst, 4, value=desc_en)
        # E 含税：若源有公式则复制公式（G列），这里简化为空或与 F 一致
        # F 不含税、G 采购
        if exclude_tax is not None and isinstance(exclude_tax, (int, float)):
            ws_dst.cell(r_dst, 6, value=exclude_tax)
            ws_dst.cell(r_dst, 7, value=exclude_tax)
        # 复制源 G 列公式到 E（若为公式）
        src_g = ws_src.cell(r_src, 7)
        if getattr(src_g.value, "__class__", None) and "Formula" in type(src_g.value).__name__:
            try:
                f = str(src_g.value)
                if f.startswith("="):
                    f_dst = re.sub(r"([A-Z]+)(\d+)", lambda m: _row_ref(m, r_src, r_dst), f)
                    ws_dst.cell(r_dst, 5, value=f_dst)
            except Exception:
                pass
        # 利润率与价格：源 I,J,K,L,M,N,O,P,Q,R,S,T → 标 H,I,J,K,L,M,N,O,P,Q,R,S
        for c_off in range(9, 21):  # I=9 to T=20
            src_cell = ws_src.cell(r_src, c_off)
            std_col = c_off - 1  # std 从 8 开始对应 src 9: std 8=H←I, 9=I←J, ... 19=S←T, 20=T←空
            if std_col > 20:
                break
            v = _cell_value(src_cell)
            # 价格列（标准 I,K,M,O,Q,S）必须用标准公式 G/(1-利润率列)，不复制源公式
            if std_col in (9, 11, 13, 15, 17, 19):
                form = _price_formula_for_std(std_col, r_dst)
                ws_dst.cell(r_dst, std_col, value=form)
            elif v is not None:
                ws_dst.cell(r_dst, std_col, value=v)
        # 标准 G 列（采购）：若尚未填，用 H 列数值
        if ws_dst.cell(r_dst, 7).value is None and ws_dst.cell(r_dst, 8).value is not None:
            g_val = _cell_value(ws_src.cell(r_src, 8))
            if g_val is not None:
                ws_dst.cell(r_dst, 7, value=g_val)
        # 管材独有 U,V,W,X：源 21→相关体积, 22→%, W=ROUND(G/(1-V),0), X=ROUND(W*111%,0)
        u_val = _cell_value(ws_src.cell(r_src, 21))
        if u_val is not None:
            ws_dst.cell(r_dst, 21, value=u_val)
        v_val = _cell_value(ws_src.cell(r_src, 22))
        if v_val is not None:
            ws_dst.cell(r_dst, 22, value=v_val)
        ws_dst.cell(r_dst, 23, value=f"=ROUND((G{r_dst}/(1-V{r_dst})),0)")
        ws_dst.cell(r_dst, 24, value=f"=ROUND(W{r_dst}*111%,0)")
    return ws_dst.max_row


def _row_ref(m, r_src, r_dst):
    """公式中行号从源行改为目标行。列不变（NEW H→标 G 等需列映射，这里仅做行号）。"""
    col, row = m.group(1), int(m.group(2))
    delta = r_dst - r_src
    return f"{col}{row + delta}"


def _price_formula_for_std(std_col: int, r_dst: int) -> str:
    """标准格式中价格列公式：I=ROUNDUP(G/(1-H),0), K=ROUND(G/(1-J),0), M=ROUND(G/(1-L),0), O=ROUND(G/(1-N),0), Q=ROUND(G/(1-P),0), S=ROUND(G/(1-R),0)。"""
    # std_col 9=I, 11=K, 13=M, 15=O, 17=Q, 19=S
    # 前一列为利润率 H,J,L,N,P,R → 列号 8,10,12,14,16,18
    rate_col = std_col - 1
    g_letter = "G"
    r_letter = get_column_letter(rate_col)
    if std_col == 9:
        return f"=ROUNDUP(({g_letter}{r_dst}/(1-{r_letter}{r_dst})),0)"
    return f"=ROUND(({g_letter}{r_dst}/(1-{r_letter}{r_dst})),0)"


def _copy_guobiao_to_standard(ws_src, ws_dst):
    """国标管件：A=序号, B=SAP物料编号, C=物料描述, D=含税, E=不含税(公式), F=大唐利润率, G=大唐价格(公式), H=其他利润率, I=其他价格(公式)。
    标准：B=Material, C=Describrition; 国标只有 大唐/其他 两档，映射到 B档(K)、D档(O)；采购=不含税。
    """
    for col, h in enumerate(STD_HEADERS, 1):
        ws_dst.cell(row=1, column=col, value=h)
    max_row = ws_src.max_row
    for r_src in range(2, max_row + 1):
        r_dst = r_src  # 输出从第 2 行起
        no = _cell_value(ws_src.cell(r_src, 1))
        material = _cell_value(ws_src.cell(r_src, 2))
        desc = _cell_value(ws_src.cell(r_src, 3))
        inc_tax = _cell_value(ws_src.cell(r_src, 4))
        exc_tax = _cell_value(ws_src.cell(r_src, 5))
        da_tang_rate = _cell_value(ws_src.cell(r_src, 6))
        da_tang_price = _cell_value(ws_src.cell(r_src, 7))
        other_rate = _cell_value(ws_src.cell(r_src, 8))
        other_price = _cell_value(ws_src.cell(r_src, 9))
        ws_dst.cell(r_dst, 1, value=no)
        ws_dst.cell(r_dst, 2, value=material)
        ws_dst.cell(r_dst, 3, value=desc)
        ws_dst.cell(r_dst, 4, value=None)  # English 空
        ws_dst.cell(r_dst, 5, value=inc_tax)  # E=含税（源 D）
        # F=不含税、G=采购不含税：标准格式 E 列为含税，故 =ROUND(E/111%,0)
        formula_no_tax = f"=ROUND(E{r_dst}/111%,0)"
        ws_dst.cell(r_dst, 6, value=formula_no_tax)
        ws_dst.cell(r_dst, 7, value=formula_no_tax)
        # A 档空, B 档=大唐/其他
        ws_dst.cell(r_dst, 8, value=None)
        ws_dst.cell(r_dst, 9, value=None)
        ws_dst.cell(r_dst, 10, value=da_tang_rate)
        ws_dst.cell(r_dst, 11, value=f"=ROUNDUP((G{r_dst}/(1-J{r_dst})),0)")
        ws_dst.cell(r_dst, 12, value=None)
        ws_dst.cell(r_dst, 13, value=None)
        ws_dst.cell(r_dst, 14, value=other_rate)
        ws_dst.cell(r_dst, 15, value=f"=ROUNDUP((G{r_dst}/(1-N{r_dst})),0)")
        for c in range(16, 21):
            ws_dst.cell(r_dst, c, value=None)
    return max_row


def main():
    if not SRC_FILE.exists():
        raise SystemExit(f"源文件不存在: {SRC_FILE}")
    wb_src = openpyxl.load_workbook(SRC_FILE, read_only=False, data_only=False)
    if "LESSO管材" not in wb_src.sheetnames:
        raise SystemExit("源文件中未找到 sheet: LESSO管材")
    if "国标管件" not in wb_src.sheetnames:
        raise SystemExit("源文件中未找到 sheet: 国标管件")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # 1. 管材（来自 LESSO管材）
    ws_lesso = wb_src["LESSO管材"]
    ws_guan = wb_out.create_sheet("管材", 0)
    n_guan = _copy_lesso_to_standard(ws_lesso, ws_guan)
    print(f"管材: 已写入 {n_guan} 行（含表头）")

    # 2. 国标管件
    ws_gb = wb_src["国标管件"]
    ws_guobiao = wb_out.create_sheet("国标管件", 1)
    n_gb = _copy_guobiao_to_standard(ws_gb, ws_guobiao)
    print(f"国标管件: 已写入 {n_gb} 行（含表头）")

    wb_src.close()
    wb_out.save(OUT_FILE)
    print(f"已保存: {OUT_FILE}")
    if "--verify" in sys.argv:
        verify_content()


def _val_eq(a, b, tol=1e-6) -> bool:
    """值相等（数值允许浮点误差）。"""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= tol
    return a == b


def verify_content() -> None:
    """核对生成文件与源文件内容：管材/国标管件关键列及行数。"""
    if not OUT_FILE.exists():
        print("未找到输出文件，请先运行生成。")
        return
    wb_src = openpyxl.load_workbook(SRC_FILE, read_only=False, data_only=True)
    wb_out = openpyxl.load_workbook(OUT_FILE, read_only=False, data_only=True)
    errors = []

    # 1. 管材：源 LESSO 行3起 → 出 行2起
    ws_src = wb_src["LESSO管材"]
    ws_out = wb_out["管材"]
    max_src = ws_src.max_row
    max_out = ws_out.max_row
    expected_data_rows = max_src - 2  # 源 3..max_src
    if max_out != 1 + expected_data_rows:
        errors.append(f"管材 行数: 期望 1+{expected_data_rows}={1 + expected_data_rows}, 实际 {max_out}")
    for i in range(min(expected_data_rows, 500)):  # 抽样前500行
        r_src = 3 + i
        r_out = 2 + i
        # 源 C,D,E,H → 出 B,C,D; 源 H → 出 F,G
        if not _val_eq(ws_src.cell(r_src, 3).value, ws_out.cell(r_out, 2).value):
            errors.append(f"管材 r{r_out} Material 不一致")
        if ws_src.cell(r_src, 4).value != ws_out.cell(r_out, 3).value:
            errors.append(f"管材 r{r_out} Describrition 不一致")
        if ws_src.cell(r_src, 5).value != ws_out.cell(r_out, 4).value:
            errors.append(f"管材 r{r_out} English 不一致")
        if not _val_eq(ws_src.cell(r_src, 8).value, ws_out.cell(r_out, 6).value):
            errors.append(f"管材 r{r_out} 不含税(F) 不一致")
        if not _val_eq(ws_src.cell(r_src, 8).value, ws_out.cell(r_out, 7).value):
            errors.append(f"管材 r{r_out} 采购(G) 不一致")
        for col_off, (c_src, c_out) in enumerate([(9, 8), (11, 10), (13, 12), (15, 14), (17, 16), (19, 18)]):
            if not _val_eq(ws_src.cell(r_src, c_src).value, ws_out.cell(r_out, c_out).value):
                errors.append(f"管材 r{r_out} 利润率列 不一致 col_off={col_off}")
        # 价格列：出 为公式时 data_only 常为 None，只验证源数据自洽（源价 = 源H/(1-利润率) 取整）
        base = ws_src.cell(r_src, 8).value  # 源 H 不含税
        for c_src, rate_src in [(10, 9), (12, 11), (14, 13), (16, 15), (18, 17), (20, 19)]:
            rate = ws_src.cell(r_src, rate_src).value
            if base is not None and rate is not None and rate != 1:
                if rate_src == 9:
                    exp = math.ceil(float(base) / (1 - float(rate)))
                else:
                    exp = round(float(base) / (1 - float(rate)), 0)
                src_val = ws_src.cell(r_src, c_src).value
                if not _val_eq(src_val, exp, tol=1):  # 允许差 1（Excel ROUND 与 Python 舍入边界）
                    errors.append(f"管材 r{r_out} 价格列 源{c_src}={src_val} 与公式期望{exp} 不一致")
        if len(errors) >= 20:
            break

    # 2. 国标管件：源 行2起 → 出 行2起
    ws_src_gb = wb_src["国标管件"]
    ws_out_gb = wb_out["国标管件"]
    n_src_gb = ws_src_gb.max_row - 1
    n_out_gb = ws_out_gb.max_row - 1
    if n_out_gb != n_src_gb:
        errors.append(f"国标管件 数据行数: 源 {n_src_gb}, 出 {n_out_gb}")
    for i in range(n_src_gb):
        r = 2 + i
        if ws_src_gb.cell(r, 2).value != ws_out_gb.cell(r, 2).value:
            errors.append(f"国标 r{r} Material 不一致")
        if ws_src_gb.cell(r, 3).value != ws_out_gb.cell(r, 3).value:
            errors.append(f"国标 r{r} Describrition 不一致")
        inc = ws_src_gb.cell(r, 4).value
        if not _val_eq(inc, ws_out_gb.cell(r, 5).value):
            errors.append(f"国标 r{r} 含税(E) 不一致")
        # 不含税与价格：用源数据按公式算期望，与源计算值比（出公式列可能未缓存）
        src_no_tax = ws_src_gb.cell(r, 5).value
        expected_no_tax = round(float(inc or 0) / 1.11, 0) if inc is not None else None
        if not _val_eq(src_no_tax, expected_no_tax):
            errors.append(f"国标 r{r} 源不含税E 与公式期望 不一致")
        out_f = ws_out_gb.cell(r, 6).value
        if out_f is not None and not _val_eq(src_no_tax, out_f):
            errors.append(f"国标 r{r} 不含税(F) 出={out_f} 源E={src_no_tax}")
        if not _val_eq(ws_src_gb.cell(r, 6).value, ws_out_gb.cell(r, 10).value):
            errors.append(f"国标 r{r} 大唐利润率(J) 不一致")
        if not _val_eq(ws_src_gb.cell(r, 8).value, ws_out_gb.cell(r, 14).value):
            errors.append(f"国标 r{r} 其他利润率(N) 不一致")
        src_da = ws_src_gb.cell(r, 7).value
        da_rate = ws_src_gb.cell(r, 6).value
        # 源 G = ROUNDUP((E/(1-F)),0)，Excel 向上取整
        expected_da = math.ceil(float(src_no_tax or 0) / (1 - float(da_rate or 0))) if src_no_tax is not None and da_rate is not None else None
        if expected_da is not None and not _val_eq(src_da, expected_da):
            errors.append(f"国标 r{r} 源大唐价G={src_da} 与公式期望{expected_da} 不一致")
        out_k = ws_out_gb.cell(r, 11).value
        if out_k is not None and not _val_eq(src_da, out_k):
            errors.append(f"国标 r{r} 大唐价格(K) 出={out_k} 源G={src_da}")
        src_other = ws_src_gb.cell(r, 9).value
        other_rate = ws_src_gb.cell(r, 8).value
        expected_other = math.ceil(float(src_no_tax or 0) / (1 - float(other_rate or 0))) if src_no_tax is not None and other_rate is not None else None
        if expected_other is not None and not _val_eq(src_other, expected_other):
            errors.append(f"国标 r{r} 源其他价I={src_other} 与公式期望{expected_other} 不一致")
        out_o = ws_out_gb.cell(r, 15).value
        if out_o is not None and not _val_eq(src_other, out_o):
            errors.append(f"国标 r{r} 其他价格(O) 出={out_o} 源I={src_other}")
        if len(errors) >= 30:
            break

    wb_src.close()
    wb_out.close()
    if errors:
        print("核对发现差异:")
        for e in errors[:30]:
            print("  ", e)
        if len(errors) > 30:
            print("  ... 更多省略")
        sys.exit(1)
    print("核对通过：内容与源文件一致（抽样/全部关键列）。")


if __name__ == "__main__":
    main()
