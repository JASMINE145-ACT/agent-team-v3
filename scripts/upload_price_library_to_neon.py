"""
将 Excel 价格库上传到 Neon，覆盖现有数据并新增 Product_Type 列。

用法：
    python scripts/upload_price_library_to_neon.py
"""

from __future__ import annotations

import io
from pathlib import Path

import psycopg2
from openpyxl import load_workbook

DATABASE_URL = "postgresql://neondb_owner:npg_WUz1AKf4OPnL@ep-shiny-frost-a16b6ixd-pooler.ap-southeast-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"
TABLE_NAME = "万鼎价格库_管材与国标管件_标准格式"
EXCEL_PATH = Path("data/万鼎价格库_管材与国标管件_标准格式.xlsx")
BATCH_SIZE = 500

# Excel 列顺序（0-based）→ Neon 列名
EXCEL_INDEX_TO_NEON_COL: dict[int, str] = {
    0: "NO",
    1: "Material",
    2: "Describrition",
    3: "Describrition_English",
    4: "Product_Type",
    5: "INCLUDE_TAX_出厂价_含税",
    6: "EXCLUDE_TAX_出厂价_不含税",
    7: "采购不含税",
    8: "（二级代理）A级别_利润率",
    9: "（二级代理）A级别_报单价格",
    10: "（一级代理）B级别_利润率",
    11: "（一级代理）B级别_报单价格",
    12: "（聚万大客户）C级别_利润率",
    13: "（聚万大客户）C级别报单价格",
    14: "（青山大客户）D级别_利润率",
    15: "（青山大客户）D级别_报单价格",
    16: "（青山大客户）D级别_降低利润率",
    17: "（青山大客户）D级别_报单价格_2",
    18: "（大唐大客户）E级别（包运费）_利润率",
    19: "（大唐大客户）E级别包运费）_报单价格",
    20: "col_20",  # Excel null 列 → Neon col_20
    21: "相关体积",
    22: "%",
    23: "EXC_TAX",
    24: "INC_TAX",
}


def q(name: str) -> str:
    """双引号包裹 SQL 标识符"""
    return '"' + name.replace('"', '""') + '"'


def read_excel(path: Path) -> tuple[list, list[list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return list(rows[0]), [list(rows[i]) for i in range(1, len(rows))]


def build_csv_lines(
    data: list[list],
    neon_cols: list[str],
) -> tuple[list[str], int, int]:
    """将 Excel 行数据转换为 CSV 文本行（用于 COPY）。

    每行产生 len(neon_cols) 个字段，用 \\t 分隔，空值也输出空串。
    COPY 的 NULL '' 选项会将空串转为 SQL NULL。
    跳过 Material 列非数字的行（Neon 该列为 numeric）。
    返回 (lines, skipped_count, total_count)。
    """
    # 预计算反向映射：neon_col → excel_idx
    neon_to_excel: dict[str, int] = {}
    for ex_idx, ne_col in EXCEL_INDEX_TO_NEON_COL.items():
        neon_to_excel[ne_col] = ex_idx

    lines = []
    skipped = 0
    for row in data:
        # 跳过 Material 非数字的行
        mat_idx = neon_to_excel.get("Material")
        if mat_idx is not None and mat_idx < len(row):
            mat_val = row[mat_idx]
            if mat_val is not None and mat_val != "":
                try:
                    float(str(mat_val))
                except (ValueError, TypeError):
                    skipped += 1
                    continue

        row_strs = []
        for neon_col in neon_cols:
            excel_idx = neon_to_excel.get(neon_col)
            if excel_idx is not None and excel_idx < len(row):
                val = row[excel_idx]
                if val is None or val == "":
                    val_str = ""
                else:
                    val_str = str(val)
                    val_str = val_str.replace("\t", " ").replace("\n", " ").replace("\r", " ")
            else:
                val_str = ""
            row_strs.append(val_str)
        lines.append("\t".join(row_strs))
    return lines, skipped, len(data)


def main() -> None:
    print(f"[*] Reading Excel: {EXCEL_PATH}")
    headers, data = read_excel(EXCEL_PATH)
    print(f"[*] Excel: {len(data)} rows, {len(headers)} cols")

    # ── Connect ────────────────────────────────────────────────
    print("[*] Connecting to Neon...")
    conn = psycopg2.connect(DATABASE_URL)
    conn.set_client_encoding("UTF8")
    cur = conn.cursor()

    # 1. ALTER TABLE: add Product_Type column
    print("[*] Adding Product_Type column...")
    try:
        cur.execute(
            f'ALTER TABLE {q(TABLE_NAME)} ADD COLUMN IF NOT EXISTS "Product_Type" TEXT'
        )
        conn.commit()
        print("  OK")
    except Exception as e:
        conn.rollback()
        print(f"  {e}")

    # 2. Get Neon columns
    cur.execute(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_name = %s ORDER BY ordinal_position",
        (TABLE_NAME,),
    )
    neon_cols = [r[0] for r in cur.fetchall()]
    print(f"[*] Neon cols: {len(neon_cols)}")

    # 3. DELETE existing data
    print("[*] Clearing existing data...")
    cur.execute(f'DELETE FROM {q(TABLE_NAME)}')
    conn.commit()
    print("  OK")

    # 4. COPY in batches (faster than INSERT, no param binding issues)
    col_list = ", ".join(q(c) for c in neon_cols)
    tbl = q(TABLE_NAME)
    copy_sql = (
        f"COPY {tbl} ({col_list}) FROM STDIN "
        "WITH (FORMAT TEXT, DELIMITER E'\\t', NULL '')"
    )

    total = len(data)
    total_skipped = 0
    for batch_start in range(0, total, BATCH_SIZE):
        batch = data[batch_start : batch_start + BATCH_SIZE]
        lines, batch_skipped, _ = build_csv_lines(batch, neon_cols)
        total_skipped += batch_skipped
        csv_content = "\n".join(lines) + "\n"
        buf = io.StringIO(csv_content)

        try:
            cur.copy_expert(copy_sql, buf)
            conn.commit()
            print(f"  Copied {min(batch_start + BATCH_SIZE, total)}/{total} (skipped: {total_skipped})")
        except Exception as e:
            conn.rollback()
            print(f"  COPY failed at {batch_start}: {e}")
            raise

    # 5. Verify
    cur.execute(f'SELECT COUNT(*) FROM {q(TABLE_NAME)}')
    count = cur.fetchone()[0]
    print(f"[*] Row count: {count}")

    cur.execute(
        f'SELECT "Product_Type", COUNT(*) FROM {q(TABLE_NAME)} '
        'WHERE "Product_Type" IS NOT NULL GROUP BY "Product_Type" LIMIT 10'
    )
    print(f"[*] Product_Type distribution: {list(cur.fetchall())}")

    cur.execute(
        f'SELECT "NO", "Material", "Describrition", "Product_Type" '
        f'FROM {q(TABLE_NAME)} ORDER BY "NO" LIMIT 3'
    )
    for r in cur.fetchall():
        print(f"  {r}")

    cur.close()
    conn.close()
    print("[*] Done!")


if __name__ == "__main__":
    main()
