#!/usr/bin/env python3
"""完整验证：规格提取 + 数据保存 + Excel 裁剪"""
import sys
import os
sys.path.insert(0, r"D:\Projects\agent-jk\Agent Team version3")

print("=" * 80)
print("测试 1：规格提取逻辑")
print("=" * 80)

from backend.tools.quotation.spec_extract import extract_spec_from_quote_name

test_cases = [
    "直通(管箍)PVC-U排水配件白色 dn50",
    "短型顺水三通印尼(日标)PVC-U管件(D排水系列)灰色 DN50 (2\") 联塑",
]

for quote_name in test_cases:
    spec = extract_spec_from_quote_name(quote_name)
    print(f"\n报价名称: {quote_name}")
    print(f"提取规格: '{spec}'")
    print(f"  - 类型: {type(spec)}")
    print(f"  - 是否为空: {spec == ''}")
    print(f"  - 布尔值: {bool(spec)}")

print("\n" + "=" * 80)
print("测试 2：数据保存逻辑")
print("=" * 80)

# 测试不同情况
test_data = [
    {"quote_spec": "PVC-U排水 dn50"},  # 有值
    {"quote_spec": ""},  # 空字符串
    {"quote_spec": None},  # None
    {},  # 缺失字段
]

for i, line in enumerate(test_data):
    print(f"\n测试 {i+1}: line = {line}")
    
    # 原逻辑（有 BUG）
    old_logic = str(line.get("quote_spec", ""))[:500] if line.get("quote_spec") else None
    
    # 新逻辑（已修复）
    new_logic = (str(line.get("quote_spec") or "")[:500]) or None
    
    print(f"  原逻辑: {repr(old_logic)}")
    print(f"  新逻辑: {repr(new_logic)}")
    
    if old_logic != new_logic:
        print(f"  ⚠️ 结果不同!")

print("\n" + "=" * 80)
print("测试 3：Excel 列裁剪")
print("=" * 80)

try:
    import openpyxl
    print("\n创建测试工作簿...")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 填充数据到 A-Z 列
    for col in range(1, 27):  # A-Z
        ws.cell(1, col, value=f"Col{col}")
    
    print(f"初始列数: {ws.max_column}")
    
    # 执行裁剪
    MAX_COL = 18  # A-R
    if ws.max_column > MAX_COL:
        ws.delete_cols(MAX_COL + 1, ws.max_column - MAX_COL)
        print(f"裁剪后列数: {ws.max_column}")
        print(f"✅ 列裁剪成功! 保留了 A-R 列")
    else:
        print(f"列数 {ws.max_column} <= {MAX_COL}, 无需裁剪")
        
except ImportError:
    print("⚠️ openpyxl 未安装")
except Exception as e:
    print(f"❌ Excel 裁剪测试失败: {e}")

print("\n" + "=" * 80)
print("测试 4：配置检查")
print("=" * 80)

try:
    from backend.config import Config
    print(f"\nQUOTATION_SPEC_LLM = {getattr(Config, 'QUOTATION_SPEC_LLM', 'NOT FOUND')}")
    print(f"OPENAI_API_KEY = {'已设置' if getattr(Config, 'OPENAI_API_KEY', None) else '未设置'}")
except Exception as e:
    print(f"配置读取失败: {e}")

print("\n" + "=" * 80)
print("测试完成")
print("=" * 80)
